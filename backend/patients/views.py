import unicodedata
import re
import uuid
import json
import ast
import os
import io
from datetime import datetime, timedelta
from http.client import RemoteDisconnected
from urllib import request as urllib_request
from urllib import error as urllib_error

import pandas as pd
from django.db import connection
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from users.permissions import IsAdminOrChefService, CanViewPatients
from rest_framework.permissions import AllowAny
from audit.models import AuditLog

from .models import Patient, PatientFormField, PatientFormTemplate
from .serializers import PatientFormTemplateSerializer, PatientSerializer
from .preprocess_rag import _env_int, build_rag_context, estimate_route


class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        if isinstance(obj, datetime):
            return obj.isoformat()
        obj_module = getattr(obj.__class__, '__module__', '')
        if obj_module.startswith('numpy'):
            if hasattr(obj, 'tolist'):
                try:
                    return obj.tolist()
                except Exception:
                    pass
            if hasattr(obj, 'item'):
                try:
                    return obj.item()
                except Exception:
                    pass
            return str(obj)
        return super().default(obj)


def _safe_json_value(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return None if value != value else value

    if isinstance(value, str):
        return value

    if isinstance(value, dict):
        return {str(key): _safe_json_value(item) for key, item in value.items()}

    if isinstance(value, (list, tuple, set)):
        return [_safe_json_value(item) for item in value]

    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()

    if isinstance(value, pd.Series):
        return [_safe_json_value(item) for item in value.tolist()]

    if isinstance(value, pd.DataFrame):
        return [_safe_json_value(row) for row in value.to_dict(orient='records')]

    value_module = getattr(value.__class__, '__module__', '')
    if value_module.startswith('numpy'):
        if hasattr(value, 'tolist'):
            try:
                return _safe_json_value(value.tolist())
            except Exception:
                pass
        if hasattr(value, 'item'):
            try:
                return _safe_json_value(value.item())
            except Exception:
                pass

    if hasattr(value, 'to_pydatetime'):
        try:
            return value.to_pydatetime().isoformat()
        except Exception:
            pass

    if hasattr(value, 'isoformat'):
        try:
            return value.isoformat()
        except Exception:
            pass

    if hasattr(value, 'to_dict'):
        try:
            return _safe_json_value(value.to_dict())
        except Exception:
            pass

    return str(value)


# ============================================================
# FONCTIONS UTILITAIRES DE BASE
# ============================================================

def normalize_header(value):
    text = unicodedata.normalize('NFKD', str(value).strip().lower())
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return text.strip('_')


def _format_patient_audit_value(value):
    if value in [None, '']:
        return '-'
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        except Exception:
            return str(value)
    return str(value)


def _describe_patient_changes(before_data, after_data, max_changes=5):
    changes = []
    before_data = before_data or {}
    after_data = after_data or {}

    keys = sorted(set(before_data.keys()) | set(after_data.keys()))
    for key in keys:
        if key in {'id', 'created_at', 'updated_at'}:
            continue
        before_value = _format_patient_audit_value(before_data.get(key))
        after_value = _format_patient_audit_value(after_data.get(key))
        if before_value == after_value:
            continue
        changes.append(f"{key}: {before_value} -> {after_value}")
        if len(changes) >= max_changes:
            break

    if not changes:
        return "aucun champ notable modifié"

    remaining = max(0, len(keys) - len(changes))
    suffix = f" (+{remaining} autre(s) champ(s))" if remaining > 0 else ''
    return '; '.join(changes) + suffix


# Plage raisonnable de serial Excel : du 01/01/1990 au 31/12/2099
_EXCEL_SERIAL_MIN = 32874   # 01/01/1990
_EXCEL_SERIAL_MAX = 73050   # 31/12/2099
_EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_serial_to_date_iso(value):
    """Convertit un serial Excel (entier) en chaîne ISO YYYY-MM-DD, ou None si hors plage."""
    try:
        ival = int(float(value))
        if _EXCEL_SERIAL_MIN <= ival <= _EXCEL_SERIAL_MAX:
            return (_EXCEL_EPOCH + timedelta(days=ival)).date().isoformat()
    except (ValueError, TypeError):
        pass
    return None


def convert_excel_value(value):
    if pd.isna(value):
        return None

    if hasattr(value, 'item') and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except Exception:
            pass

    # Objet datetime/date pandas ou Python
    if hasattr(value, 'to_pydatetime'):
        return value.to_pydatetime().date().isoformat()

    if hasattr(value, 'date') and not isinstance(value, str):
        try:
            return value.date().isoformat()
        except Exception:
            return str(value)

    # Entier ou flottant
    if isinstance(value, (int, float)):
        fval = float(value)
        if fval != fval:  # NaN
            return None
        if fval.is_integer():
            ival = int(fval)
            # Essayer la conversion serial Excel avant de retourner l'entier brut
            date_iso = excel_serial_to_date_iso(ival)
            if date_iso:
                return date_iso
            return ival
        return fval

    if hasattr(value, 'isoformat') and not isinstance(value, str):
        try:
            return value.isoformat()
        except Exception:
            pass

    # Chaîne : tenter de détecter un serial ou une date texte
    if isinstance(value, str):
        stripped = value.strip()
        # Chaîne purement numérique → tester serial Excel
        if stripped.isdigit():
            date_iso = excel_serial_to_date_iso(int(stripped))
            if date_iso:
                return date_iso
        return stripped

    return value


def normalize_type(value):
    normalized = normalize_header(value).replace('_', ' ')
    TYPE_MAP = {
        'texte libre court': 'text_short',
        'texte libre long': 'text_long',
        'liste a choix unique': 'single_choice',
        'liste a choix multiple': 'multiple_choice',
        'selecteur de date': 'date',
        'nombre entier': 'integer',
        'nombre decimal': 'decimal',
        'oui/non': 'boolean',
        'genere automatiquement': 'auto',
    }
    return TYPE_MAP.get(normalized, 'text_short')


def parse_flexible_date(value):
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    parsed = parse_date(text)
    if parsed:
        return parsed.isoformat()

    for dayfirst in [True, False]:
        try:
            dt = pd.to_datetime(text, errors='coerce', dayfirst=dayfirst)
            if pd.notna(dt):
                return dt.date().isoformat()
        except Exception:
            continue

    return None


def normalize_age_value(value):
    if value in [None, '']:
        return None

    try:
        age_value = int(float(str(value).strip()))
        if age_value < 0:
            return None
        return age_value
    except Exception:
        return None


def derive_age_from_date_of_birth(date_iso):
    if not date_iso:
        return None

    birth_date = parse_date(str(date_iso))
    if not birth_date:
        return None

    today = timezone.localdate()
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age if age >= 0 else None


def derive_date_of_birth_from_age(age_value):
    normalized_age = normalize_age_value(age_value)
    if normalized_age is None:
        return None
    target_year = timezone.localdate().year - normalized_age
    return f"{target_year}-01-01"


def normalize_sex_values(value):
    if value is None:
        return None, None

    value_str = str(value).strip().lower()

    if value_str in ['1', '1.0']:
        return 'M', 'homme'
    if value_str in ['0', '0.0']:
        return 'F', 'femme'

    normalized = normalize_header(value_str)
    if normalized in ['m', 'male', 'masculin', 'homme', 'man']:
        return 'M', 'homme'
    if normalized in ['f', 'female', 'feminin', 'femme', 'woman']:
        return 'F', 'femme'
    if normalized in ['i', 'intersex', 'intersexe']:
        return 'O', 'intersexe'
    if normalized in ['unknown', 'inconnu', 'na', 'n_a', 'none', 'null', '', '9', '9.0']:
        return 'O', 'inconnu'
    return 'O', 'inconnu'


# ============================================================
# DÉCODAGE DES VALEURS NUMÉRIQUES CODÉES
# ============================================================

# Tables de correspondance complètes (issues du fichier Excel de référence)
_DECODE_MAPS = {
    # Démographie
    'sexe': {1: 'homme', 0: 'femme'},
    'demographie_sexe': {1: 'homme', 0: 'femme'},
    'couverture_sociale': {
        0: 'auto_paiement',
        1: 'ramed',
        2: 'amo',
        3: 'autre_assurance_publique',
    },
    'demographie_couverture_sociale': {
        0: 'auto_paiement',
        1: 'ramed',
        2: 'amo',
        3: 'autre_assurance_publique',
    },

    # IRC / étiologie
    'etiologie_mrc': {
        1: 'nephropathie_diabetique',
        2: 'nephropathie_indeterminee',
        3: 'vascularite_anca',
        4: 'maladie_renale_polykystique',
        5: 'nephroangiosclerose',
        6: 'uropathie_obstructive',
        7: 'nephropathie_lupique',
        8: 'glomerulonephrite_membraneuse',
        9: 'nephropathie_a_iga',
        10: 'lgm_hsf',
        11: 'amylose_myelome',
        12: 'gn_crescentique',
        13: 'syndrome_hemolytique_et_uremique',
        14: 'glomerulopathie_c3',
        15: 'necrose_corticale',
    },
    'irc_etiologie_principale': {
        1: 'nephropathie_diabetique',
        2: 'nephropathie_indeterminee',
        3: 'vascularite_anca',
        4: 'maladie_renale_polykystique',
        5: 'nephroangiosclerose',
        6: 'uropathie_obstructive',
        7: 'nephropathie_lupique',
        8: 'glomerulonephrite_membraneuse',
        9: 'nephropathie_a_iga',
        10: 'lgm_hsf',
        11: 'amylose_myelome',
        12: 'gn_crescentique',
        13: 'syndrome_hemolytique_et_uremique',
        14: 'glomerulopathie_c3',
        15: 'necrose_corticale',
    },
    'groupe_etiologie_mrc': {
        1: 'diabete',
        2: 'nephroangiosclerose',
        3: 'polykystose',
        4: 'uropathie_obstructive',
        5: 'lupus',
        6: 'vascularite',
        7: 'autre_glomerulaire',
        8: 'autre',
        9: 'indeterminee',
    },

    # Dialyse
    'type_acces_initial': {
        1: 'cathetere_femoral',
        2: 'cathetere_tunnellise',
        3: 'fistule_arterioveineuse',
        4: 'cathetere_peritoneal',
    },
    'dialyse_type_acces_initial': {
        1: 'cathetere_femoral',
        2: 'cathetere_tunnellise',
        3: 'fistule_arterioveineuse',
        4: 'cathetere_peritoneal',
    },
    'fistule_arterioveineuse': {1: 'oui', 0: 'non'},
    'fistule_arterioveineuse_c': {1: 'oui', 0: 'non'},
    'groupe_jours_entre_cathetere': {
        0: 'pas_d_intervalle',
        1: '0_30_jours',
        2: '31_181_jours',
        3: 'plus_de_180_jours',
    },
    'dialyse_jours_entre_catheter_et_fav': {
        0: 'pas_d_intervalle',
        1: '0_30_jours',
        2: '31_181_jours',
        3: 'plus_de_180_jours',
    },

    # Statut diurèse
    'statut_diurese': {1: 'anurique', 2: 'diurese_preservee'},
    'presentation_statut_diurese': {1: 'anurique', 2: 'diurese_preservee'},

    # Devenir / décès
    'deces': {1: 'oui', 0: 'non', 9: 'inconnu'},
    'cause_deces': {
        1: 'cardiovasculaire',
        2: 'infection',
        3: 'hemorragique',
        4: 'autre',
        5: 'indeterminee',
    },
    'devenir_cause_deces': {
        1: 'cardiovasculaire',
        2: 'infection',
        3: 'hemorragique',
        4: 'autre',
        5: 'indeterminee',
    },
}

# Colonnes déjà fusionnées dans comorbidite_liste — ne doivent jamais apparaître
# comme colonnes séparées dans la structure de la plateforme
_COLUMNS_FUSED_INTO_COMORBIDITE_LISTE = {
    'exposition_toxique',
    'comorbidite_exposition_toxique',
    'antecedents_medicaments_nephrotoxiques',
    'comorbidite_antecedents_medicaments_nephrotoxiques',
}

# Colonnes booléennes simples (1 → oui, 0 → non)
_BOOLEAN_COLUMNS = {
    'hypertension', 'cardiopathie', 'hemodialyse', 'dialyse_peritoneale',
    'debut_dialyse_urgence', 'debut_dialyse_planifie',
    'cause_deces_cardiaque', 'cause_deces_infectieuse',
    'irc_maladie_renale_hereditaire', 'irc_antecedents_familiaux_renaux',
    'irc_connue_avant_dialyse',
    'dialyse_information_transplantation_donnee',
    'irc_themes_education_therapeutique',
    'biopsie_renale', 'maladie_renale_hereditaire',
    'uropathie_obstructive', 'goutte', 'exposition_toxique',
    'antecedents_medicaments_nephrotoxiques',
    'asthenie', 'douleur_abdominale', 'nausees', 'prurit',
    'asymptomatique', 'trouble_conscience',
    'infection', 'trouble_electrolytique', 'evenement_cardiovasculaire',
    'hemorragie', 'dysfonction_acces', 'crise_convulsive',
    'information_transplantation', 'information_transplantation_donnee',
    'liste_attente_transplantation', 'transplantation_renale',
    'immunisation_transfusion_sanguine', 'bilan_pretransplantation',
    'debut_dialyse_urgence', 'debut_dialyse_planifie',
    'changement_hd_vers_dp', 'changement_dp_vers_hd',
}

def is_binary_column(column_data):
    """
    Détermine si une colonne est binaire (contient uniquement 0/1 et variantes).
    column_data: liste des valeurs de la colonne (non None)
    Retourne True si toutes les valeurs non vides sont 0 ou 1 (ou 0.0/1.0)
    """
    if not column_data:
        return False
    
    for value in column_data:
        if value is None or value == '':
            continue
        
        try:
            # Convertir en float pour gérer 0.0 et 1.0
            num_value = float(value)
            if num_value not in (0.0, 1.0):
                return False
        except (ValueError, TypeError):
            # Si ce n'est pas un nombre, ce n'est pas binaire
            return False
    
    return True
def decode_numeric_value(column_name, value):
    """Convertit les valeurs numériques codées en libellés textuels."""
    if value is None or value == '':
        return value

    col = normalize_header(str(column_name))

    # Chercher dans les tables de décodage
    decode_map = _DECODE_MAPS.get(col)
    if decode_map:
        try:
            return decode_map.get(int(float(value)), value)
        except (ValueError, TypeError):
            return value

    # Colonnes booléennes simples
    if col in _BOOLEAN_COLUMNS:
        try:
            val_int = int(float(value))
            return 'oui' if val_int == 1 else 'non'
        except (ValueError, TypeError):
            return value

    # Détection générique : si la valeur est strictement 0 ou 1 (entier), convertir en oui/non
    try:
        val_str = str(value).strip()
        if val_str in ('0', '1'):
            return 'oui' if val_str == '1' else 'non'
    except Exception:
        pass

    return value


def extract_charlson_score(value):
    """Extrait le score Charlson d'une valeur potentiellement mal formatée."""
    if value is None or value == '':
        return None

    str_value = str(value).strip()

    try:
        return int(float(str_value))
    except (ValueError, TypeError):
        pass

    numbers = re.findall(r'\d+', str_value)
    if numbers:
        return int(numbers[0])

    return None


def _is_truthy(value):
    if isinstance(value, bool):
        return value
    normalized = normalize_header(value)
    return normalized in ['1', 'true', 'yes', 'oui', 'y']


def _is_falsey(value):
    if isinstance(value, bool):
        return not value
    normalized = normalize_header(value)
    return normalized in ['0', 'false', 'no', 'non', 'aucun', 'none', 'null', 'na', 'n_a', '']


def _pg_quote_identifier(value):
    return '"' + str(value).replace('"', '""') + '"'


def _pg_quote_literal(value):
    return "'" + str(value).replace("'", "''") + "'"


# ============================================================
# CHARGEMENT AUTOMATIQUE DU MAPPING DEPUIS column_mapping.json
# ============================================================

def load_column_mapping():
    """Charge le mapping depuis le fichier JSON."""
    mapping_file = os.path.join(os.path.dirname(__file__), 'column_mapping.json')
    try:
        with open(mapping_file, 'r', encoding='utf-8') as f:
            mapping_list = json.load(f)

        column_mapping = {}
        for item in mapping_list:
            main_key = normalize_header(item['main'])
            platform_key = normalize_header(item['platform'])
            transformation = item['transformation']

            if transformation == 'direct':
                transform_type = 'direct'
            elif transformation == 'calcul (age → date naissance estimée)':
                transform_type = 'calcul_age_to_birthdate'
            elif transformation in [
                'fusion (urgence/planifié → contexte)',
                'fusion (hd/dp → modalité)',
                'fusion (cardiaque/infectieux → cause)',
                'fusion (hd→dp / dp→hd → changement)',
            ]:
                transform_type = transformation
            elif 'fusion' in transformation or 'inclus dans' in transformation:
                transform_type = transformation
            else:
                transform_type = 'direct'

            column_mapping[main_key] = {
                'platform': platform_key,
                'type': transform_type,
                'original_main': item['main'],
                'original_platform': item['platform'],
            }

        return column_mapping
    except Exception as e:
        print(f"Erreur chargement mapping: {e}")
        return get_default_mapping()


def get_default_mapping():
    """Mapping par défaut — couvre les colonnes les plus fréquentes."""
    return {
        'identifiant_patient': {'platform': 'id_patient', 'type': 'direct'},
        'sexe': {'platform': 'demographie_sexe', 'type': 'direct'},
        'age_annees': {'platform': 'demographie_date_naissance', 'type': 'calcul_age_to_birthdate'},
        'distance_au_centre': {'platform': 'demographie_distance_centre_km', 'type': 'direct'},
        'etiologie_mrc': {'platform': 'irc_etiologie_principale', 'type': 'direct'},
        'charlson': {'platform': 'icc_charlson', 'type': 'direct'},
        'icc_charlson': {'platform': 'icc_charlson', 'type': 'direct'},
        # --- Dates de dialyse (variantes de nommage) ---
        'date_debut_dialyse': {'platform': 'dialyse_date_debut', 'type': 'direct'},
        'date_debut': {'platform': 'dialyse_date_debut', 'type': 'direct'},
        'dialyse_date_debut': {'platform': 'dialyse_date_debut', 'type': 'direct'},
        'date_demarrage_dialyse': {'platform': 'dialyse_date_debut', 'type': 'direct'},
        'debut_dialyse': {'platform': 'dialyse_date_debut', 'type': 'direct'},
        # --- Comorbidités ---
        'hypertension': {'platform': 'comorbidite_liste', 'type': 'fusion (valeurs multiples → liste)'},
        'cardiopathie': {'platform': 'comorbidite_liste', 'type': 'fusion (valeurs multiples → liste)'},
        # --- Dialyse modalité ---
        'hemodialyse': {'platform': 'dialyse_modalite_initiale', 'type': 'fusion (hd/dp → modalité)'},
        'dialyse_peritoneale': {'platform': 'dialyse_modalite_initiale', 'type': 'fusion (hd/dp → modalité)'},
        # --- Contexte début dialyse ---
        'debut_dialyse_urgence': {'platform': 'irc_contexte_debut_dialyse', 'type': 'fusion (urgence/planifié → contexte)'},
        'debut_dialyse_planifie': {'platform': 'irc_contexte_debut_dialyse', 'type': 'fusion (urgence/planifié → contexte)'},
        # --- Devenir ---
        'deces': {'platform': 'devenir_statut', 'type': 'inclus dans statut devenir'},
        'cause_deces_cardiaque': {'platform': 'devenir_cause_deces', 'type': 'fusion (cardiaque/infectieux → cause)'},
        'cause_deces_infectieuse': {'platform': 'devenir_cause_deces', 'type': 'fusion (cardiaque/infectieux → cause)'},
    }


# Charger le mapping au démarrage
COLUMN_MAPPING = load_column_mapping()

SECTION_PREFIX_MAP = {
    'demographie_': 'demographie_data',
    'irc_': 'irc_data',
    'comorbidite_': 'comorbidite_data',
    'presentation_': 'presentation_data',
    'biologie_': 'biologie_data',
    'imagerie_': 'imagerie_data',
    'dialyse_': 'dialyse_data',
    'qualite_': 'qualite_data',
    'complication_': 'complication_data',
    'traitement_': 'traitement_data',
    'devenir_': 'devenir_data',
}

TEMPLATE_KEYWORDS = {
    'texte libre court',
    'texte libre long',
    'liste a choix unique',
    'liste a choix multiple',
    'selecteur de date',
    'nombre entier',
    'oui/non',
    'genere automatiquement',
}

FIXED_CLASSEUR_TEMPLATE_NAMES = [
    'template',
    'template_patients_hd',
    'plateform_donnees_complete',
    'classeur1',
    'Main',
]

AUTO_INCREMENT_FIELD_PREFIX = {
    'id_patient': 'PAT',
    'id_enregistrement_source': 'SRC',
}

POSTGRES_MODEL_FIELD_MAP = {
    'id_patient': 'id_patient',
    'id_enregistrement_source': 'id_enregistrement_source',
    'id_site': 'id_site',
    'statut_inclusion': 'statut_inclusion',
    'statut_consentement': 'statut_consentement',
    'utilisateur_saisie': 'utilisateur_saisie',
    'derniere_mise_a_jour': 'derniere_mise_a_jour',
    'date_evaluation_initiale': 'date_evaluation_initiale',
    'nom': 'nom',
    'prenom': 'prenom',
    'age': 'age',
    'sexe': 'sexe',
    'maladie': 'maladie',
    'telephone': 'telephone',
    'adresse': 'adresse',
    'date_naissance': 'date_naissance',
    'date_admission': 'date_admission',
}

POSTGRES_SECTION_BUCKETS = [
    ('demographie_', 'demographie_data'),
    ('irc_', 'irc_data'),
    ('comorbidite_', 'comorbidite_data'),
    ('presentation_', 'presentation_data'),
    ('biologie_', 'biologie_data'),
    ('imagerie_', 'imagerie_data'),
    ('dialyse_', 'dialyse_data'),
    ('qualite_', 'qualite_data'),
    ('complication_', 'complication_data'),
    ('traitement_', 'traitement_data'),
    ('devenir_', 'devenir_data'),
]


def refresh_postgres_flat_view(template=None):
    if template is None:
        template = PatientFormTemplate.objects.filter(name__iexact='template').order_by('-id').first()
        if template is None:
            template = PatientFormTemplate.objects.order_by('-id').first()

    if template is None:
        return

    keys = [
        key for key in template.fields.order_by('order', 'id').values_list('key', flat=True)
        if key and not key.startswith('unnamed')
    ]
    if not keys:
        return

    fixed_keys = [
        'id_patient', 'nom', 'prenom', 'age', 'sexe', 'maladie',
        'telephone', 'adresse', 'date_naissance', 'date_admission',
        'id_enregistrement_source', 'id_site', 'statut_inclusion',
        'statut_consentement', 'date_evaluation_initiale',
        'utilisateur_saisie', 'derniere_mise_a_jour',
    ]

    select_parts = ['p.id AS id']
    for key in fixed_keys:
        if key in POSTGRES_MODEL_FIELD_MAP:
            select_parts.append(f"p.{POSTGRES_MODEL_FIELD_MAP[key]} AS {_pg_quote_identifier(key)}")

    for key in keys:
        if key in fixed_keys:
            continue
        if key in POSTGRES_MODEL_FIELD_MAP:
            expr = f"p.{POSTGRES_MODEL_FIELD_MAP[key]}"
        else:
            bucket = None
            for prefix, column_name in POSTGRES_SECTION_BUCKETS:
                if key.startswith(prefix):
                    bucket = column_name
                    break

            if bucket:
                expr = f"p.{bucket} ->> {_pg_quote_literal(key)}"
            else:
                expr = f"p.extra_data ->> {_pg_quote_literal(key)}"

        select_parts.append(f"{expr} AS {_pg_quote_identifier(key)}")

    sql_create = (
        'CREATE VIEW public.patients_plateforme_flat AS '
        'SELECT ' + ', '.join(select_parts) + ' FROM patients_patient p'
    )

    with connection.cursor() as cursor:
        # DROP obligatoire : CREATE OR REPLACE ne peut pas renommer des colonnes existantes
        cursor.execute('DROP VIEW IF EXISTS public.patients_plateforme_flat CASCADE')
        cursor.execute(sql_create)


def get_active_template():
    template = None
    for template_name in FIXED_CLASSEUR_TEMPLATE_NAMES:
        template = (
            PatientFormTemplate.objects.filter(name__iexact=template_name).order_by('-id').first()
            or PatientFormTemplate.objects.filter(sheet_name__iexact=template_name).order_by('-id').first()
        )
        if template:
            break
    if not template:
        template = PatientFormTemplate.objects.order_by('-id').first()
    return template


def is_schema_template_sheet(worksheet):
    if worksheet.max_row < 2:
        return False

    type_cells = [worksheet.cell(2, col_index).value for col_index in range(1, worksheet.max_column + 1)]
    normalized_types = {
        normalize_header(cell).replace('_', ' ')
        for cell in type_cells
        if cell is not None and str(cell).strip()
    }
    return bool(normalized_types.intersection(TEMPLATE_KEYWORDS))


def create_or_update_template(worksheet, source_file_name):
    template_name = (worksheet.title if worksheet else None) or 'Patient Schema'
    template, _ = PatientFormTemplate.objects.update_or_create(
        name=template_name,
        defaults={
            'source_file_name': source_file_name or template_name,
            'sheet_name': worksheet.title if worksheet else template_name,
        },
    )
    return template


def parse_schema_from_template_sheet(worksheet, source_file_name):
    template = create_or_update_template(worksheet, source_file_name)
    template.fields.all().delete()

    created_fields = []
    for column_index in range(1, worksheet.max_column + 1):
        header = worksheet.cell(1, column_index).value
        type_label = worksheet.cell(2, column_index).value
        if not header:
            continue

        choices = []
        for row_index in range(3, worksheet.max_row + 1):
            value = worksheet.cell(row_index, column_index).value
            if value is not None and str(value).strip():
                choices.append(str(value).strip())

        field_type = normalize_type(type_label)
        created_fields.append(
            PatientFormField(
                template=template,
                key=normalize_header(header),
                label=str(header).strip(),
                field_type=field_type,
                order=column_index,
                choices=choices,
                source_hint=str(type_label or ''),
                is_required=field_type not in ['auto'],
            )
        )

    PatientFormField.objects.bulk_create(created_fields)
    try:
        refresh_postgres_flat_view(template)
    except Exception:
        pass
    return template, len(created_fields)


def upsert_template_from_headers(headers, worksheet, source_file_name, create_fields=True):
    template = None
    for template_name in FIXED_CLASSEUR_TEMPLATE_NAMES:
        template = (
            PatientFormTemplate.objects.filter(name__iexact=template_name).order_by('-id').first()
            or PatientFormTemplate.objects.filter(sheet_name__iexact=template_name).order_by('-id').first()
        )
        if template:
            break

    if template is None:
        template = create_or_update_template(worksheet, source_file_name)
    else:
        if source_file_name:
            template.source_file_name = source_file_name
        if worksheet and getattr(worksheet, 'title', None):
            template.sheet_name = worksheet.title
        template.save(update_fields=['source_file_name', 'sheet_name'])

    if not create_fields:
        try:
            refresh_postgres_flat_view(template)
        except Exception:
            pass
        return template

    existing_fields = {field.key: field for field in template.fields.all()}

    for index, header in enumerate(headers, start=1):
        if not header:
            continue

        header_str = str(header).strip()
        if not header_str or header_str.lower().startswith('unnamed'):
            continue

        key = normalize_header(header_str)
        if not key or key.startswith('unnamed') or key in existing_fields:
            continue

        # Ne jamais créer de champ pour les colonnes fusionnées dans comorbidite_liste
        if key in _COLUMNS_FUSED_INTO_COMORBIDITE_LISTE:
            continue

        PatientFormField.objects.create(
            template=template,
            key=key,
            label=str(header).strip(),
            field_type='text_short',
            order=index,
            choices=[],
            source_hint='auto_detected_from_data_import',
            is_required=False,
        )

    try:
        refresh_postgres_flat_view(template)
    except Exception:
        pass

    return template


def build_patient_payload(row):
    # DEBUG: afficher les colonnes une seule fois
    if not hasattr(build_patient_payload, '_cols_printed'):
        print("=== COLONNES DU FICHIER EXCEL ===")
        for i, col in enumerate(row.keys()):
            print(f'  {i}: "{col}"')
        build_patient_payload._cols_printed = True

    payload = {'extra_data': {}}

    # Initialiser les structures par section
    payload.setdefault('demographie_data', {})
    payload.setdefault('irc_data', {})
    payload.setdefault('comorbidite_data', {})
    payload.setdefault('presentation_data', {})
    payload.setdefault('biologie_data', {})
    payload.setdefault('dialyse_data', {})
    payload.setdefault('complication_data', {})
    payload.setdefault('devenir_data', {})

    comorbidite_list = []
    presentation_list = []
    complication_list = []

    contexte_dialyse = None
    modalite_dialyse = None
    date_transplantation = None
    cause_deces_value = None
    deces_statut = None
    changement_modalite = False
    age_value = None
    age_column_found = False

    # PARCOURIR TOUTES LES COLONNES
    for column_name, raw_value in row.items():
        value = convert_excel_value(raw_value)
        if value is None:
            continue

        # Décodage numérique AVANT toute autre logique
        value = decode_numeric_value(column_name, value)
        normalized = normalize_header(column_name)

        # === DÉTECTION DE L'ÂGE ===
        if not age_column_found and (
            'age' in normalized
            or normalized in ('ans', 'years', 'age_annees')
        ):
            age_value = normalize_age_value(value)
            if age_value is not None and 0 < age_value < 120:
                payload['age'] = age_value
                payload['demographie_data']['demographie_age_ans'] = age_value
                date_estimee = derive_date_of_birth_from_age(age_value)
                if date_estimee:
                    payload['demographie_data']['demographie_date_naissance'] = date_estimee
                    # Marquer que la date est estimée (calculée depuis l'âge, pas saisie réelle)
                    payload['extra_data']['demographie_date_naissance_estimee'] = True
                age_column_found = True
                print(f"Âge détecté: '{column_name}' = {age_value} ans")
            continue

        # === DÉTECTION CHARLSON ===
        col_name_lower = str(column_name).lower().strip()
        is_charlson = (
            col_name_lower in ['charlson', 'icc_charlson', 'icccharlson', 'score_charlson']
            or str(column_name).strip() in ['Charlson', 'ICC_Charlson', 'ICC_CHARLSON']
            or 'charlson' in col_name_lower
        )

        if is_charlson:
            charlson_val = extract_charlson_score(value)
            if charlson_val is not None:
                payload['icc_charlson'] = str(charlson_val)
                print(f"✅ CHARLSON CAPTURÉ: '{column_name}' → {charlson_val}")
            else:
                print(f"⚠️ Colonne Charlson trouvée mais valeur non extraite: '{raw_value}'")

        # === COLONNES DÉJÀ FUSIONNÉES DANS comorbidite_liste → IGNORER ===
        # (traitées via le mapping fusion, ne doivent pas créer de colonne séparée)
        if normalized in _COLUMNS_FUSED_INTO_COMORBIDITE_LISTE:
            # La fusion est gérée plus bas via COLUMN_MAPPING si la valeur est truthy
            pass

        # === MAPPING STANDARD ===
        mapping = COLUMN_MAPPING.get(normalized)

        if mapping:
            platform_field = mapping['platform']
            transform_type = mapping['type']

            if transform_type == 'direct':
                section_target = None
                for prefix, section_bucket in SECTION_PREFIX_MAP.items():
                    if platform_field.startswith(prefix):
                        section_target = section_bucket
                        break

                if section_target:
                    payload[section_target][platform_field] = value
                else:
                    payload[platform_field] = value

            elif transform_type == 'calcul_age_to_birthdate':
                if not age_column_found:
                    age_value = normalize_age_value(value)
                    if age_value is not None and 0 < age_value < 120:
                        payload['age'] = age_value
                        payload['demographie_data']['demographie_age_ans'] = age_value
                        date_estimee = derive_date_of_birth_from_age(age_value)
                        if date_estimee:
                            payload['demographie_data']['demographie_date_naissance'] = date_estimee
                            # Marquer que la date est estimée (calculée depuis l'âge, pas saisie réelle)
                            payload['extra_data']['demographie_date_naissance_estimee'] = True
                        age_column_found = True

            elif transform_type == 'fusion (urgence/planifié → contexte)':
                if _is_truthy(value):
                    contexte_dialyse = 'debut_en_urgence' if 'urgence' in normalized else 'debut_planifie'

            elif transform_type == 'fusion (hd/dp → modalité)':
                if _is_truthy(value):
                    modalite_dialyse = 'hemodialyse' if 'hemodialyse' in normalized else 'dialyse_peritoneale'

            elif transform_type == 'partiel (oui/non → date)':
                if _is_truthy(value):
                    date_transplantation = timezone.now().date().isoformat()

            elif transform_type == 'fusion (valeurs multiples → liste)' and platform_field == 'comorbidite_liste':
                if _is_truthy(value):
                    label = normalized.replace('_', ' ')
                    if label not in comorbidite_list:
                        comorbidite_list.append(label)
                        print(f"Comorbidité: {label}")

            elif 'symptome' in transform_type.lower() or (
                transform_type == 'fusion (valeurs multiples → liste)'
                and platform_field == 'presentation_symptomes'
            ):
                if _is_truthy(value):
                    label = normalized.replace('_', ' ')
                    if label not in presentation_list:
                        presentation_list.append(label)

            elif 'complication' in transform_type.lower() or (
                transform_type == 'fusion (valeurs multiples → liste)'
                and platform_field == 'complication_liste'
            ):
                if _is_truthy(value):
                    label = normalized.replace('_', ' ')
                    if label not in complication_list:
                        complication_list.append(label)

            elif 'inclus dans thèmes éducation' in transform_type:
                # Ces colonnes sont maintenant ignorées et iront dans extra_data
                # car elles n'ont plus d'entrée dans column_mapping.json
                # On ne fait rien ici, elles seront capturées par le else
                pass

            elif transform_type == 'fusion (cardiaque/infectieux → cause)':
                if _is_truthy(value):
                    if 'cardiaque' in normalized:
                        cause_deces_value = 'cardiovasculaire'
                    elif 'infectieuse' in normalized:
                        cause_deces_value = 'infection'

            elif transform_type == 'inclus dans statut devenir' and platform_field == 'devenir_statut':
                deces_statut = value

            elif 'changement' in transform_type.lower():
                if _is_truthy(value):
                    changement_modalite = True
        else:
            # Colonnes déjà fusionnées → ne jamais créer de colonne séparée
            if normalized in _COLUMNS_FUSED_INTO_COMORBIDITE_LISTE:
                pass
            else:
                # Colonnes non mappées → extra_data (devient une colonne dynamique)
                payload['extra_data'][normalized] = value
                # Mémoriser les noms de colonnes dynamiques pour le rapport d'import
                payload.setdefault('_dynamic_columns_detected', set()).add(normalized)
                print(f"🆕 Colonne dynamique ajoutée: '{column_name}' = {value}")

    # ── Appliquer les fusions ──────────────────────────────────────────────
    if comorbidite_list:
        payload['comorbidite_data']['comorbidite_liste'] = ', '.join(comorbidite_list)

    if presentation_list:
        payload['presentation_data']['presentation_symptomes'] = ', '.join(presentation_list)

    if complication_list:
        payload['complication_data']['complication_liste'] = ', '.join(complication_list)

    if contexte_dialyse:
        payload['irc_data']['irc_contexte_debut_dialyse'] = contexte_dialyse

    if modalite_dialyse:
        payload['dialyse_data']['dialyse_modalite_initiale'] = modalite_dialyse

    if date_transplantation:
        payload['devenir_data']['devenir_date_transplantation'] = date_transplantation

    if cause_deces_value:
        payload['devenir_data']['devenir_cause_deces'] = cause_deces_value

    if deces_statut is not None:
        payload['devenir_data']['devenir_statut'] = (
            'decede' if _is_truthy(deces_statut) else 'vivant_sous_dialyse'
        )

    if changement_modalite:
        payload['complication_data']['complication_changement_modalite_dialyse'] = 'oui'

    # ── Normalisation du sexe ─────────────────────────────────────────────
    if 'demographie_sexe' in payload['demographie_data']:
        dem_sexe = payload['demographie_data']['demographie_sexe']
        patient_sex, demo_sex = normalize_sex_values(dem_sexe)
        if patient_sex:
            payload['sexe'] = patient_sex
        if demo_sex:
            payload['demographie_data']['demographie_sexe'] = demo_sex

    # ── Récapitulatif ────────────────────────────────────────────────────
    if age_column_found:
        print(f"✓ Âge: {age_value} ans")

    if 'icc_charlson' in payload:
        print(f"✅ CHARLSON DANS LE PAYLOAD: {payload['icc_charlson']}")
    else:
        print("❌ Score Charlson NON trouvé dans ce fichier")

    return payload


def ensure_required_identity_fields(payload):
    extra_data = payload.get('extra_data') or {}
    identifier = payload.get('id_patient') or extra_data.get('id_patient')

    if not payload.get('id_patient') and identifier:
        payload['id_patient'] = str(identifier)

    nom_value = payload.get('nom')
    if nom_value is None or str(nom_value).strip() == '':
        payload['nom'] = 'Import_Automatique'

    prenom_value = payload.get('prenom')
    if prenom_value is None or str(prenom_value).strip() == '':
        payload['prenom'] = f"Patient_{payload.get('id_patient', '000')}"

    payload['extra_data'] = extra_data
    return payload


def _extract_numeric_suffix(value):
    """Extrait le dernier bloc de chiffres d'une chaîne (après le dernier tiret ou caractère non-numérique)."""
    if value is None:
        return None
    text = str(value).strip()
    
    # Chercher le dernier bloc de chiffres consécutifs
    digits = ''
    for char in reversed(text):
        if char.isdigit():
            digits = char + digits
        elif digits:
            # On a trouvé un bloc de chiffres, arrêter
            break
    
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def generate_next_incremental_identifier(model_field, prefix):
    max_number = 0
    for current_value in Patient.objects.values_list(model_field, flat=True):
        numeric_value = _extract_numeric_suffix(current_value)
        if numeric_value and numeric_value > max_number:
            max_number = numeric_value
    next_number = max_number + 1
    return f"{prefix}{next_number:06d}"


def generate_next_numeric_patient_id(auto_increment_state=None):
    if auto_increment_state is not None:
        auto_increment_state['id_patient'] = auto_increment_state.get('id_patient', 0) + 1
        return str(auto_increment_state['id_patient'])
    max_number = 0
    for current_value in Patient.objects.values_list('id_patient', flat=True):
        numeric_value = _extract_numeric_suffix(current_value)
        if numeric_value and numeric_value > max_number:
            max_number = numeric_value
    return str(max_number + 1)


def initialize_auto_increment_state():
    state = {}
    for model_field in AUTO_INCREMENT_FIELD_PREFIX:
        max_number = 0
        for current_value in Patient.objects.values_list(model_field, flat=True):
            numeric_value = _extract_numeric_suffix(current_value)
            if numeric_value and numeric_value > max_number:
                max_number = numeric_value
        state[model_field] = max_number
    return state


def resolve_entry_user_label(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return 'system_import'
    for attr in ['username', 'email']:
        value = getattr(user, attr, None)
        if value:
            return str(value)
    return str(getattr(user, 'id', 'system_import'))


def ensure_incremental_identifiers(payload, auto_increment_state=None, force_generated=False):
    if force_generated:
        patient_id = generate_next_numeric_patient_id(auto_increment_state=auto_increment_state)
        payload['id_patient'] = patient_id
        
        # Générer id_enregistrement_source basé sur id_patient pour assurer la synchronisation
        patient_num = _extract_numeric_suffix(patient_id)
        if patient_num is not None:
            payload['id_enregistrement_source'] = (
                f"{AUTO_INCREMENT_FIELD_PREFIX['id_enregistrement_source']}{patient_num:06d}"
            )
        else:
            # Fallback si extraction échoue
            payload['id_enregistrement_source'] = generate_next_incremental_identifier(
                'id_enregistrement_source',
                AUTO_INCREMENT_FIELD_PREFIX['id_enregistrement_source'],
            )
        return payload
    if not payload.get('id_patient'):
        payload['id_patient'] = generate_next_numeric_patient_id(auto_increment_state=auto_increment_state)
    return payload


def apply_automatic_schema_fields(payload, auto_increment_state=None, current_user=None):
    template = PatientFormTemplate.objects.order_by('-id').first()
    if not template:
        return payload
    extra_data = payload.get('extra_data') or {}
    for field in template.fields.filter(field_type='auto'):
        if field.key in ['id_patient', 'id_enregistrement_source']:
            continue
    payload['extra_data'] = extra_data

    if 'utilisateur_saisie' not in payload:
        payload['utilisateur_saisie'] = resolve_entry_user_label(current_user)
    if 'derniere_mise_a_jour' not in payload:
        payload['derniere_mise_a_jour'] = timezone.now().isoformat()
    if 'date_evaluation_initiale' not in payload:
        payload['date_evaluation_initiale'] = timezone.now().date().isoformat()

    return payload


PREPROCESS_SESSION_DIR = os.path.join(os.path.dirname(__file__), 'preprocess_sessions')
CRITICAL_PREPROCESS_KEYS = ['nom', 'prenom', 'id_patient']


def _ensure_preprocess_session_dir():
    os.makedirs(PREPROCESS_SESSION_DIR, exist_ok=True)


def _preprocess_session_path(session_id):
    safe_id = re.sub(r'[^a-zA-Z0-9_-]', '', str(session_id or ''))
    return os.path.join(PREPROCESS_SESSION_DIR, f'{safe_id}.json')


def _save_preprocess_session(session_payload):
    _ensure_preprocess_session_dir()
    session_id = session_payload.get('id') or uuid.uuid4().hex
    session_payload['id'] = session_id
    session_payload['updated_at'] = timezone.now().isoformat()
    session_path = _preprocess_session_path(session_id)
    with open(session_path, 'w', encoding='utf-8') as handle:
        json.dump(_safe_json_value(session_payload), handle, ensure_ascii=False)
    return session_payload


def _load_preprocess_session(session_id):
    session_path = _preprocess_session_path(session_id)
    if not os.path.exists(session_path):
        return None
    with open(session_path, 'r', encoding='utf-8') as handle:
        return json.load(handle)


def _to_json_compatible(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if isinstance(value, (dict, list, str, int, float, bool)):
        return value
    return str(value)


def _dataframe_to_rows(dataframe):
    rows = []
    for _, row in dataframe.iterrows():
        payload = {}
        for column in dataframe.columns:
            payload[str(column)] = _to_json_compatible(row[column])
        rows.append(payload)
    return rows


def _rows_to_dataframe(rows, columns=None):
    if not rows:
        return pd.DataFrame(columns=columns or [])
    dataframe = pd.DataFrame(rows)
    if columns:
        for column in columns:
            if column not in dataframe.columns:
                dataframe[column] = None
        dataframe = dataframe[columns]
    return dataframe


def _read_uploaded_dataframe(uploaded_file):
    source_file_name = getattr(uploaded_file, 'name', 'uploaded_file')
    lower_name = str(source_file_name).lower()

    if lower_name.endswith('.csv'):
        uploaded_file.seek(0)
        dataframe = pd.read_csv(uploaded_file)
    elif lower_name.endswith('.xls'):
        uploaded_file.seek(0)
        dataframe = pd.read_excel(uploaded_file, engine='xlrd')
    else:
        uploaded_file.seek(0)
        dataframe = pd.read_excel(uploaded_file, parse_dates=True)

    dataframe = dataframe.loc[:, ~dataframe.columns.astype(str).str.match(r'^(Unnamed|unnamed)(:.*)?$')]
    return dataframe, source_file_name


def _build_technical_profile(dataframe):
    columns_profile = []
    numeric_columns_profile = []
    categorical_columns_profile = []
    total_rows = int(len(dataframe.index))
    total_columns = int(len(dataframe.columns))
    total_cells = total_rows * total_columns
    missing_cells = int(dataframe.isna().sum().sum()) if total_cells else 0
    duplicate_rows = int(dataframe.duplicated().sum()) if total_rows else 0
    duplicate_ratio = round((duplicate_rows / total_rows) * 100, 2) if total_rows else 0.0

    for column in dataframe.columns:
        series = dataframe[column]
        non_null_series = series.dropna()
        missing_count = int(series.isna().sum())
        non_null_count = int(series.notna().sum())
        sample_values = []
        for value in non_null_series.astype(str).head(3).tolist():
            if value not in sample_values:
                sample_values.append(value)

        column_profile = {
            'column': str(column),
            'dtype': str(series.dtype),
            'non_null_count': non_null_count,
            'missing_count': missing_count,
            'missing_pct': round((missing_count / total_rows) * 100, 2) if total_rows else 0.0,
            'sample_values': sample_values,
        }
        columns_profile.append(column_profile)

        if pd.api.types.is_numeric_dtype(series):
            numeric_series = pd.to_numeric(series, errors='coerce').dropna()
            if not numeric_series.empty:
                q1 = float(numeric_series.quantile(0.25))
                q3 = float(numeric_series.quantile(0.75))
                iqr = q3 - q1
                lower_bound = q1 - (1.5 * iqr)
                upper_bound = q3 + (1.5 * iqr)
                outlier_count = int(((numeric_series < lower_bound) | (numeric_series > upper_bound)).sum())
                numeric_columns_profile.append({
                    'column': str(column),
                    'count': int(numeric_series.count()),
                    'mean': round(float(numeric_series.mean()), 4),
                    'std': round(float(numeric_series.std(ddof=0)), 4) if numeric_series.count() > 1 else 0.0,
                    'min': round(float(numeric_series.min()), 4),
                    'q1': round(q1, 4),
                    'median': round(float(numeric_series.median()), 4),
                    'q3': round(q3, 4),
                    'max': round(float(numeric_series.max()), 4),
                    'outlier_count': outlier_count,
                })
        else:
            top_values = non_null_series.astype(str).value_counts().head(5)
            categorical_columns_profile.append({
                'column': str(column),
                'unique_count': int(non_null_series.astype(str).nunique()),
                'top_values': [
                    {'value': str(index), 'count': int(count)}
                    for index, count in top_values.items()
                ],
            })

    return {
        'rows': total_rows,
        'columns': total_columns,
        'total_cells': total_cells,
        'missing_cells': missing_cells,
        'missing_pct': round((missing_cells / total_cells) * 100, 2) if total_cells else 0.0,
        'duplicate_rows': duplicate_rows,
        'duplicate_pct': duplicate_ratio,
        'column_names': [str(column) for column in dataframe.columns.tolist()],
        'columns_profile': columns_profile,
        'numeric_columns_profile': numeric_columns_profile,
        'categorical_columns_profile': categorical_columns_profile,
        'preview_rows': _dataframe_to_rows(dataframe.head(3)),
    }


def _get_section_prefix_for_column(column_name):
    normalized = normalize_header(column_name)
    for prefix, section_name in SECTION_PREFIX_MAP.items():
        if normalized.startswith(prefix):
            return section_name
    return 'generic_data'


def _build_preprocess_chunks(dataframe, technical_profile, max_rows_per_chunk=40):
    """
    Build semantic chunks for LLM analysis.

    Goals:
    - Chunk by sections (column groups) using SECTION_PREFIX_MAP.
    - Preserve patient coherence when a patient identifier column is present.
    - Limit chunk size by rows and by estimated characters (env: CHUNK_MAX_CHARS).
    - Expose chunk metadata useful for later merge/merge-intelligent logic.
    """
    chunks = []
    total_rows = int(len(dataframe.index))
    columns = [str(column) for column in dataframe.columns.tolist()]

    # Configurable limits
    try:
        max_chars = int(os.environ.get('CHUNK_MAX_CHARS', '4000'))
    except Exception:
        max_chars = 4000
    try:
        max_rows = int(os.environ.get('CHUNK_MAX_ROWS', str(max_rows_per_chunk)))
    except Exception:
        max_rows = max_rows_per_chunk

    # Heuristic: detect patient identifier column to keep patient rows together
    patient_id_col = None
    id_candidates = [c for c in columns if any(token in c.lower() for token in ('patient', 'patient_id', 'id_patient', 'id', 'nid', 'no_patient', 'numero'))]
    if id_candidates:
        # prefer exact patient_id-like names
        for cand in id_candidates:
            try:
                unique_count = int(dataframe[cand].nunique(dropna=True))
                if unique_count > 0 and unique_count < max(2, total_rows // 2):
                    patient_id_col = cand
                    break
            except Exception:
                continue

    # Group columns by section prefix
    grouped_columns = {}
    for column in columns:
        grouped_columns.setdefault(_get_section_prefix_for_column(column), []).append(column)

    def _estimate_row_chars(row_series):
        try:
            return sum(len(str(v or '')) for v in row_series.tolist())
        except Exception:
            return 0

    for section_name, section_columns in grouped_columns.items():
        if not section_columns:
            continue
        section_frame = dataframe[section_columns]
        row_indices = list(section_frame.index)

        if patient_id_col and patient_id_col in dataframe.columns:
            # Build groups of contiguous rows per patient id to preserve coherence
            groups = []
            current_group = {'start': None, 'end': None, 'rows': [], 'chars': 0, 'patient_id': None}
            for idx in row_indices:
                pid = dataframe.at[idx, patient_id_col] if patient_id_col in dataframe.columns else None
                if current_group['patient_id'] is None:
                    current_group['patient_id'] = pid
                    current_group['start'] = idx
                if pid != current_group['patient_id'] and current_group['rows']:
                    current_group['end'] = current_group['rows'][-1]
                    groups.append(current_group)
                    current_group = {'start': idx, 'end': None, 'rows': [], 'chars': 0, 'patient_id': pid}
                # add row
                row_series = section_frame.loc[idx] if idx in section_frame.index else section_frame.iloc[0:0]
                rchars = _estimate_row_chars(row_series)
                current_group['rows'].append(idx)
                current_group['chars'] += rchars
            if current_group['rows']:
                current_group['end'] = current_group['rows'][-1]
                groups.append(current_group)

            # Build chunks by aggregating patient groups until limits reached
            current_chunk = None
            for grp in groups:
                if current_chunk is None:
                    current_chunk = {'rows': [], 'chars': 0, 'start': grp['start'], 'end': grp['end']}
                # If adding this group would exceed limits and current_chunk not empty, flush
                if (len(current_chunk['rows']) + len(grp['rows']) > max_rows) or (current_chunk['chars'] + grp['chars'] > max_chars and current_chunk['rows']):
                    # flush
                    start_idx = current_chunk['start']
                    end_idx = current_chunk['end']
                    chunk_frame = section_frame.loc[current_chunk['rows']]
                    chunks.append({
                        'chunk_id': f'{section_name}:rows:{start_idx}-{end_idx}',
                        'kind': 'row_batch',
                        'section': section_name,
                        'columns': section_columns,
                        'rows_range': [int(start_idx), int(end_idx)],
                        'row_count': int(len(current_chunk['rows'])),
                        'preview_rows': _dataframe_to_rows(chunk_frame.head(3)),
                        'estimated_chars': int(current_chunk['chars']),
                        'patient_id_column': patient_id_col,
                    })
                    current_chunk = {'rows': [], 'chars': 0, 'start': grp['start'], 'end': grp['end']}

                # Add group to current chunk (may exceed limits if a single patient is huge)
                current_chunk['rows'].extend(grp['rows'])
                current_chunk['chars'] += grp['chars']
                current_chunk['end'] = grp['end']

            if current_chunk and current_chunk['rows']:
                start_idx = current_chunk['start']
                end_idx = current_chunk['end']
                chunk_frame = section_frame.loc[current_chunk['rows']]
                chunks.append({
                    'chunk_id': f'{section_name}:rows:{start_idx}-{end_idx}',
                    'kind': 'row_batch',
                    'section': section_name,
                    'columns': section_columns,
                    'rows_range': [int(start_idx), int(end_idx)],
                    'row_count': int(len(current_chunk['rows'])),
                    'preview_rows': _dataframe_to_rows(chunk_frame.head(3)),
                    'estimated_chars': int(current_chunk['chars']),
                    'patient_id_column': patient_id_col,
                })
        else:
            # No patient id: simple row-based chunking with char-limit enforcement
            start_ptr = 0
            indices = row_indices
            n = len(indices)
            while start_ptr < n:
                end_ptr = min(start_ptr + max_rows, n)
                chunk_indices = indices[start_ptr:end_ptr]
                # tighten if estimated chars exceed max_chars
                chars = 0
                for i, idx in enumerate(chunk_indices):
                    row_series = section_frame.loc[idx]
                    chars += _estimate_row_chars(row_series)
                    if chars > max_chars:
                        # cut here
                        chunk_indices = chunk_indices[:i+1]
                        end_ptr = start_ptr + i + 1
                        break

                start_idx = chunk_indices[0]
                end_idx = chunk_indices[-1]
                chunk_frame = section_frame.loc[chunk_indices]
                chunks.append({
                    'chunk_id': f'{section_name}:rows:{start_idx}-{end_idx}',
                    'kind': 'row_batch',
                    'section': section_name,
                    'columns': section_columns,
                    'rows_range': [int(start_idx), int(end_idx)],
                    'row_count': int(len(chunk_frame.index)),
                    'preview_rows': _dataframe_to_rows(chunk_frame.head(3)),
                    'estimated_chars': int(chars),
                })
                start_ptr = end_ptr

    if not chunks:
        chunks.append({
            'chunk_id': 'generic:empty',
            'kind': 'fallback',
            'section': 'generic_data',
            'columns': columns,
            'rows_range': [0, 0],
            'row_count': 0,
            'preview_rows': [],
            'estimated_chars': 0,
        })

    # Attach chunk summary into technical_profile for visibility
    technical_profile['chunk_count'] = len(chunks)
    technical_profile['chunks'] = [
        {
            'chunk_id': chunk['chunk_id'],
            'kind': chunk.get('kind'),
            'section': chunk.get('section'),
            'row_count': chunk.get('row_count'),
            'columns_count': len(chunk.get('columns', [])),
            'estimated_chars': int(chunk.get('estimated_chars') or 0),
        }
        for chunk in chunks
    ]
    return chunks


def _build_retrieval_context(dataframe, chunks, technical_profile, stage_name='diagnostic', max_chunks=4, progress_callback=None):
    return build_rag_context(
        dataframe,
        chunks,
        technical_profile,
        stage_name=stage_name,
        max_chunks=max_chunks,
        progress_callback=progress_callback,
    )


def _determine_preprocess_route(technical_profile):
    return estimate_route(technical_profile)


def _strip_json_wrappers(response_text):
    text = str(response_text or '').strip()
    if text.startswith('```'):
        text = re.sub(r'^```(?:json)?\s*', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\s*```\s*$', '', text)
    return text.strip()


def _extract_balanced_json_candidate(response_text):
    text = str(response_text or '')
    start_index = None
    stack = []
    in_string = False
    escape_next = False

    for index, character in enumerate(text):
        if start_index is None:
            if character in '{[':
                start_index = index
                stack.append(character)
            continue

        if in_string:
            if escape_next:
                escape_next = False
                continue
            if character == '\\':
                escape_next = True
                continue
            if character == '"':
                in_string = False
            continue

        if character == '"':
            in_string = True
            continue

        if character in '{[':
            stack.append(character)
            continue

        if character in '}]' and stack:
            opening = stack[-1]
            if (opening == '{' and character == '}') or (opening == '[' and character == ']'):
                stack.pop()
                if not stack:
                    return text[start_index:index + 1]

    return None


def _repair_json_text(response_text):
    text = _strip_json_wrappers(response_text)
    if not text:
        return None

    # Helper: extract a candidate starting at pos (returns substring and max depth)
    def _extract_from(text, start_pos):
        stack = []
        in_string = False
        escape = False
        max_depth = 0
        for i in range(start_pos, len(text)):
            ch = text[i]
            if in_string:
                if escape:
                    escape = False
                elif ch == '\\':
                    escape = True
                elif ch == '"':
                    in_string = False
                continue
            if ch == '"':
                in_string = True
                continue
            if ch in '{[':
                stack.append(ch)
                max_depth = max(max_depth, len(stack))
                continue
            if ch in '}]' and stack:
                opening = stack[-1]
                if (opening == '{' and ch == '}') or (opening == '[' and ch == ']'):
                    stack.pop()
                    if not stack:
                        return text[start_pos:i + 1], max_depth, True
        # not balanced: return until end with depth info
        return text[start_pos:], max_depth, False

    candidates = []
    # full text as first candidate
    candidates.append({'text': text, 'source': 'full', 'depth': 0, 'balanced': False})

    # find all '{' or '[' positions and extract candidates
    for idx, ch in enumerate(text):
        if ch in '{[':
            sub, depth, balanced = _extract_from(text, idx)
            candidates.append({'text': sub.strip(), 'source': f'pos_{idx}', 'depth': depth, 'balanced': balanced})

    # De-duplicate by text
    seen = set()
    uniq_candidates = []
    for c in candidates:
        t = c['text']
        if not t or t in seen:
            continue
        seen.add(t)
        uniq_candidates.append(c)

    best_result = None
    best_score = -1.0

    original_len = len(re.sub(r"\s+","", text)) or 1

    def try_parse(s):
        try:
            return json.loads(s)
        except Exception:
            return None

    for c in uniq_candidates:
        cand = c['text']
        # 1) try direct parse
        parsed = try_parse(cand)
        if parsed is not None:
            plen = len(re.sub(r"\s+","", cand))
            base_score = min(1.0, plen / original_len)
            method = 'direct_parse'
            is_partial = (plen < original_len)
            depth_factor = min(1.0, float(c.get('depth', 0)) / 10.0)
            # critical keys presence
            CRITICAL_KEYS = ('dataset_summary', 'medical_analysis', 'corrections_applied')
            keys_found = 0
            if isinstance(parsed, dict):
                for k in CRITICAL_KEYS:
                    if k in parsed and parsed[k]:
                        keys_found += 1
            keys_score = keys_found / max(1, len(CRITICAL_KEYS))
            # final weighted score
            score = round(min(1.0, 0.6 * base_score + 0.25 * depth_factor + 0.15 * keys_score), 3)
            meta = {'recovery_score': score, 'method_used': method, 'is_partial': is_partial, 'depth': c.get('depth', 0), 'keys_found': keys_found}
            best_result = (parsed, meta) if score > best_score else best_result
            best_score = max(best_score, score)
            if score == 1.0:
                break

        # 2) try removing trailing fragments (trim) progressively at sensible breakpoints
        trimmed = None
        trim_points = [m.start() for m in re.finditer(r'[\}\]\",]', cand)]
        # include full length as last resort
        trim_points.append(len(cand) - 1)
        # iterate trimming to the last sensible point
        for tp in reversed(trim_points):
            if tp < max(10, len(cand) // 10):
                break
            sub = cand[:tp + 1]
            sub = re.sub(r',\s*([}\]])', r'\1', sub)
            parsed = try_parse(sub)
            if parsed is not None:
                plen = len(re.sub(r"\s+","", sub))
                score = min(1.0, plen / original_len)
                method = 'trim_trailing'
                is_partial = True
                meta = {'recovery_score': round(score, 3), 'method_used': method, 'is_partial': is_partial}
                if score > best_score:
                    best_result = (parsed, meta)
                    best_score = score
                trimmed = True
                break

        if trimmed:
            continue

        # 3) try auto-closing based on unmatched openers
        opens = cand.count('{') + cand.count('[')
        closes = cand.count('}') + cand.count(']')
        needed = opens - closes
        if needed > 0:
            # produce closers by scanning from end to find which types remain
            # simple heuristic: close '}' for each '{' and ']' for each '[' in occurrence order
            # compute stack of unmatched openers
            stack = []
            in_string = False
            escape = False
            for ch in cand:
                if in_string:
                    if escape:
                        escape = False
                    elif ch == '\\':
                        escape = True
                    elif ch == '"':
                        in_string = False
                    continue
                if ch == '"':
                    in_string = True
                    continue
                if ch in '{[':
                    stack.append(ch)
                elif ch in '}]' and stack:
                    if (stack[-1] == '{' and ch == '}') or (stack[-1] == '[' and ch == ']'):
                        stack.pop()
            closers = []
            while stack:
                opener = stack.pop()
                closers.append('}' if opener == '{' else ']')
            attempt = cand + ''.join(closers)
            attempt = re.sub(r',\s*([}\]])', r'\1', attempt)
            parsed = try_parse(attempt)
            if parsed is not None:
                plen = len(re.sub(r"\s+","", attempt))
                base_score = min(1.0, plen / original_len)
                depth_factor = min(1.0, float(c.get('depth', 0)) / 10.0)
                keys_found = 0
                if isinstance(parsed, dict):
                    for k in ('dataset_summary', 'medical_analysis', 'corrections_applied'):
                        if k in parsed and parsed[k]:
                            keys_found += 1
                keys_score = keys_found / 3.0
                score = round(min(1.0, 0.6 * base_score + 0.25 * depth_factor + 0.15 * keys_score), 3)
                method = 'auto_close'
                is_partial = True
                meta = {'recovery_score': score, 'method_used': method, 'is_partial': is_partial, 'depth': c.get('depth', 0), 'keys_found': keys_found}
                if score > best_score:
                    best_result = (parsed, meta)
                    best_score = score
                continue

        # 4) attempt python literal fallback (single quotes, None/True/False)
        try:
            py_cand = re.sub(r"\bnull\b", 'None', cand, flags=re.IGNORECASE)
            py_cand = re.sub(r"\btrue\b", 'True', py_cand, flags=re.IGNORECASE)
            py_cand = re.sub(r"\bfalse\b", 'False', py_cand, flags=re.IGNORECASE)
            parsed_py = ast.literal_eval(py_cand)
            if parsed_py is not None:
                plen = len(re.sub(r"\s+","", cand))
                base_score = min(1.0, plen / original_len)
                depth_factor = min(1.0, float(c.get('depth', 0)) / 10.0)
                keys_found = 0
                if isinstance(parsed_py, dict):
                    for k in ('dataset_summary', 'medical_analysis', 'corrections_applied'):
                        if k in parsed_py and parsed_py[k]:
                            keys_found += 1
                keys_score = keys_found / 3.0
                score = round(min(1.0, 0.6 * base_score + 0.25 * depth_factor + 0.15 * keys_score), 3)
                method = 'python_literal'
                is_partial = True
                meta = {'recovery_score': score, 'method_used': method, 'is_partial': is_partial, 'depth': c.get('depth', 0), 'keys_found': keys_found}
                if score > best_score:
                    best_result = (parsed_py, meta)
                    best_score = score
        except Exception:
            pass

    # If we found a result, normalize return to dict and attach metadata
    if best_result is not None:
        parsed_obj, meta = best_result

        # Determine structure type
        if isinstance(parsed_obj, list):
            structure_type = 'array_root'
        elif isinstance(parsed_obj, dict):
            structure_type = 'object'
        else:
            structure_type = 'mixed'

        # Compute domain score with presence + completeness checks
        CRITICAL_KEYS = ('dataset_summary', 'medical_analysis', 'corrections_applied')
        presence_score = 0.0
        completeness_score = 0.0

        def _dataset_summary_complete(ds):
            if not isinstance(ds, dict):
                return 0.0
            # require at least one of these fields to be present and non-empty
            for key in ('rows', 'columns', 'missing_cells', 'missing_pct'):
                if key in ds and ds[key] not in (None, {}, [], ''):
                    return 1.0
            return 0.0

        def _medical_analysis_complete(ma):
            if not isinstance(ma, dict):
                return 0.0
            # prefer explicit detected anomalies or a non-trivial summary
            issues = ma.get('issues')
            if isinstance(issues, (list, tuple)) and len(issues) > 0:
                return 1.0
            summary = ma.get('summary') or ''
            try:
                if isinstance(summary, str) and len(summary.strip()) >= 20:
                    return 1.0
            except Exception:
                pass
            # fallback: any other non-empty key indicates some analysis
            for k, v in ma.items():
                if k in ('issues', 'summary'):
                    continue
                if v not in (None, {}, [], ''):
                    return 1.0
            return 0.0

        def _corrections_applied_complete(ca):
            # must be explicitly present and a list (empty list acceptable)
            return 1.0 if isinstance(ca, list) else 0.0

        if isinstance(parsed_obj, dict):
            # presence: count keys that exist (regardless of truthiness)
            present = 0
            for k in CRITICAL_KEYS:
                if k in parsed_obj:
                    present += 1
            presence_score = present / float(len(CRITICAL_KEYS))

            # completeness per-key
            ds_comp = _dataset_summary_complete(parsed_obj.get('dataset_summary'))
            ma_comp = _medical_analysis_complete(parsed_obj.get('medical_analysis'))
            ca_comp = _corrections_applied_complete(parsed_obj.get('corrections_applied'))
            completeness_score = (ds_comp + ma_comp + ca_comp) / float(len(CRITICAL_KEYS))

        elif isinstance(parsed_obj, list):
            # for array roots, evaluate items that are dicts
            items = [it for it in parsed_obj if isinstance(it, dict)]
            if not items:
                presence_score = 0.0
                completeness_score = 0.0
            else:
                pres_values = []
                comp_values = []
                for it in items:
                    present = 0
                    for k in CRITICAL_KEYS:
                        if k in it:
                            present += 1
                    pres_values.append(present / float(len(CRITICAL_KEYS)))

                    ds_comp = _dataset_summary_complete(it.get('dataset_summary'))
                    ma_comp = _medical_analysis_complete(it.get('medical_analysis'))
                    ca_comp = _corrections_applied_complete(it.get('corrections_applied'))
                    comp_values.append((ds_comp + ma_comp + ca_comp) / float(len(CRITICAL_KEYS)))

                presence_score = sum(pres_values) / len(pres_values)
                completeness_score = sum(comp_values) / len(comp_values)

        # combine presence and completeness into domain_score
        presence_score = round(float(presence_score), 3)
        completeness_score = round(float(completeness_score), 3)
        domain_score = round(min(1.0, 0.5 * presence_score + 0.5 * completeness_score), 3)

        # Compute structure score
        depth_factor = min(1.0, float(meta.get('depth', 0)) / 10.0)
        parsed_valid = 1.0 if isinstance(parsed_obj, (dict, list)) else 0.0
        # penalize cases where method indicates no recovery
        method_used = meta.get('method_used', '') or ''
        if 'no_recovery' in method_used:
            parsed_valid = 0.0
        structure_score = round(min(1.0, 0.7 * parsed_valid + 0.3 * depth_factor), 3)

        # Final combined score (structure/domain split)
        domain_score = round(float(domain_score), 3)
        final_score = round(min(1.0, 0.4 * structure_score + 0.6 * domain_score), 3)

        # Domain gate: stricter medical rule
        # enforce a minimum domain_score threshold for trusting outputs
        domain_gate = domain_score >= 0.4

        # Annotate meta with presence/completeness breakdown
        meta.update({
            'structure_type': structure_type,
            'structure_score': structure_score,
            'domain_score': domain_score,
            'presence_score': presence_score,
            'completeness_score': completeness_score,
            'recovery_score': final_score,
            'domain_gate': domain_gate,
        })

        # set failure_type: hard if domain gate not satisfied
        meta['failure_type'] = 'hard' if not domain_gate else 'soft'

        # special handling for arrays: wrap results but require domain inspection
        if structure_type == 'array_root':
            wrapped = {'results': parsed_obj}
            # annotate method
            meta['method_used'] = meta.get('method_used', '') + '|array_root_wrapped'
            wrapped.update(meta)
            wrapped['trusted'] = bool(domain_gate)
            return wrapped

        # ensure dict and inject missing critical keys if needed
        if isinstance(parsed_obj, dict):
            injected = False
            if 'dataset_summary' not in parsed_obj:
                parsed_obj['dataset_summary'] = {}
                injected = True
            if 'medical_analysis' not in parsed_obj:
                parsed_obj['medical_analysis'] = {}
                injected = True
            if 'corrections_applied' not in parsed_obj:
                parsed_obj['corrections_applied'] = []
                injected = True
            if injected:
                meta['method_used'] = meta.get('method_used', '') + '|structure_injected'
                meta['is_partial'] = True
            parsed_obj.update(meta)
            # If domain gate failed, mark not trusted explicitly
            if not domain_gate:
                parsed_obj['trusted'] = False
            else:
                parsed_obj['trusted'] = True
            return parsed_obj

        # fallback non-dict non-list
        out = {'recovered_non_dict': parsed_obj}
        out.update(meta)
        return out

    # Nothing worked: return structured minimal fallback with low score
    excerpt = (text[:1000] + '...') if len(text) > 1000 else text
    return {
        'recovery_status': 'failed_partial_parse',
        'raw_excerpt': excerpt,
        'reason': 'truncated_llm_output',
        'recovery_score': 0.0,
        'method_used': 'no_recovery',
        'is_partial': True,
        'failure_type': 'hard',
        'domain_gate': False,
        'trusted': False,
    }


def _parse_llm_analysis_response(raw_response):
    def _default_preprocess_llm_output():
        return {
            'dataset_summary': {},
            'medical_analysis': {},
            'missing_values_analysis': {},
            'outliers_analysis': {},
            'duplicate_analysis': {},
            'corrections_applied': [],
            'suspect_values': [],
            'remaining_risks': [],
            'recommendations': [],
            'cleaned_dataset_preview': [],
            'processing_statistics': {},
            'quality_score': {},
            'summary': '',
            'issues': [],
            'correction_plan': {},
            'corrected_preview_rows': [],
            'column_assessment': [],
            'limitations': [],
            'raw_response': None,
        }

    if isinstance(raw_response, dict):
        merged = _default_preprocess_llm_output()
        merged.update(raw_response)
        return merged

    response_text = str(raw_response or '').strip()
    if not response_text:
        default_output = _default_preprocess_llm_output()
        default_output['summary'] = 'Le modele n a retourne aucun contenu exploitable.'
        default_output['limitations'] = ['Reponse vide du modele.']
        return default_output

    parsed = _repair_json_text(response_text)
    if parsed is not None:
        merged = _default_preprocess_llm_output()
        merged.update(parsed)
        merged['raw_response'] = response_text
        return merged

    # Dernier recours: tenter de récupérer uniquement la première structure JSON fermée.
    extracted = _extract_balanced_json_candidate(response_text)
    if extracted:
        parsed = _repair_json_text(extracted)
        if parsed is not None:
            merged = _default_preprocess_llm_output()
            merged.update(parsed)
            return merged

    fallback_output = _default_preprocess_llm_output()
    fallback_output['summary'] = response_text[:1200]
    fallback_output['limitations'] = ['Le modele a repondu, mais le JSON est invalide.']
    fallback_output['raw_response'] = response_text
    return fallback_output


def _build_llm_payload(dataframe, technical_profile):
    max_preview_rows = int(os.environ.get('OLLAMA_PREVIEW_ROWS', '3'))
    max_column_samples = int(os.environ.get('OLLAMA_COLUMN_SAMPLES', '1'))
    max_columns = int(os.environ.get('OLLAMA_MAX_COLUMNS', '40'))
    max_value_chars = int(os.environ.get('OLLAMA_MAX_VALUE_CHARS', '50'))

    def _llm_compact_value(value):
        normalized = _to_json_compatible(value)
        if isinstance(normalized, str) and len(normalized) > max_value_chars:
            return normalized[:max_value_chars] + '...'
        return normalized

    selected_columns = [str(column) for column in list(dataframe.columns)[:max_columns]]
    selected_df = dataframe[selected_columns] if selected_columns else dataframe

    preview_rows = _dataframe_to_rows(selected_df.head(max_preview_rows))
    for row in preview_rows:
        if isinstance(row, dict):
            for key in list(row.keys()):
                row[key] = _llm_compact_value(row.get(key))

    column_samples = {
        str(column): [
            _llm_compact_value(value)
            for value in selected_df[column].dropna().head(max_column_samples).tolist()
        ]
        for column in selected_df.columns
    }

    compact_columns_profile = []
    for column_meta in technical_profile.get('columns_profile', []):
        column_name = str(column_meta.get('name', ''))
        if column_name not in selected_columns:
            continue
        compact_columns_profile.append({
            'name': column_name,
            'dtype': column_meta.get('dtype'),
            'missing_count': column_meta.get('missing_count'),
            'missing_ratio': column_meta.get('missing_ratio'),
            'unique_values': column_meta.get('unique_values'),
        })

    compact_profile = {
        'rows': technical_profile.get('rows'),
        'columns': technical_profile.get('columns'),
        'missing_cells': technical_profile.get('missing_cells'),
        'missing_pct': technical_profile.get('missing_pct'),
        'duplicate_rows': technical_profile.get('duplicate_rows'),
        'duplicate_pct': technical_profile.get('duplicate_pct'),
        'columns_profile': compact_columns_profile,
        'numeric_columns_profile': technical_profile.get('numeric_columns_profile', [])[:12],
        'categorical_columns_profile': technical_profile.get('categorical_columns_profile', [])[:12],
        'preview_rows': technical_profile.get('preview_rows', [])[:max_preview_rows],
    }

    return {
        'pack_type': 'structured_analysis_pack',
        'technical_profile': compact_profile,
        'preview_rows': preview_rows,
        'column_samples': column_samples,
        'meta': {
            'selected_columns_count': len(selected_columns),
            'total_columns_count': int(len(dataframe.columns)),
            'preview_rows_count': len(preview_rows),
            'column_samples_per_column': max_column_samples,
        },
    }


def _shrink_llm_payload(payload, max_columns=20, max_preview_rows=2, max_samples_per_column=1):
    if not isinstance(payload, dict):
        return payload

    shrunk = json.loads(json.dumps(payload, default=str))

    technical_profile = shrunk.get('technical_profile') or {}
    columns_profile = technical_profile.get('columns_profile') or []
    selected_column_names = [
        str(item.get('name'))
        for item in columns_profile[:max_columns]
        if isinstance(item, dict) and item.get('name')
    ]

    technical_profile['columns_profile'] = columns_profile[:max_columns]
    technical_profile['numeric_columns_profile'] = (technical_profile.get('numeric_columns_profile') or [])[:max_columns]
    technical_profile['categorical_columns_profile'] = (technical_profile.get('categorical_columns_profile') or [])[:max_columns]
    technical_profile['preview_rows'] = (technical_profile.get('preview_rows') or [])[:max_preview_rows]

    preview_rows = shrunk.get('preview_rows') or []
    shrunk['preview_rows'] = preview_rows[:max_preview_rows]

    column_samples = shrunk.get('column_samples') or {}
    if selected_column_names:
        filtered_samples = {
            name: (column_samples.get(name) or [])[:max_samples_per_column]
            for name in selected_column_names
        }
    else:
        filtered_samples = {
            str(name): (values or [])[:max_samples_per_column]
            for name, values in list(column_samples.items())[:max_columns]
        }
    shrunk['column_samples'] = filtered_samples

    meta = shrunk.get('meta') or {}
    meta['selected_columns_count'] = len(filtered_samples)
    meta['preview_rows_count'] = len(shrunk['preview_rows'])
    meta['column_samples_per_column'] = max_samples_per_column
    shrunk['meta'] = meta

    return shrunk


def _estimate_tokens_from_text(text):
    try:
        s = str(text or '')
        return max(1, int(len(s) / 4))
    except Exception:
        return 1


def _estimate_prompt_and_rag_tokens(payload_for_prompt, retrieval_context):
    try:
        payload_text = json.dumps(payload_for_prompt, ensure_ascii=False, default=str)
    except Exception:
        payload_text = str(payload_for_prompt)
    try:
        rag_text = json.dumps(retrieval_context or {}, ensure_ascii=False, default=str)
    except Exception:
        rag_text = str(retrieval_context)
    return _estimate_tokens_from_text(payload_text), _estimate_tokens_from_text(rag_text)


def _shrink_prompt_for_budget(payload_for_prompt, retrieval_context, available_tokens):
    """
    Reduce prompt payload and retrieval_context to fit within available_tokens.
    Reduction order (priority): history -> retrieved_chunks (RAG secondary) -> column_samples -> preview_rows -> descriptive text
    """
    # Work on copies to avoid mutating caller unexpectedly
    try:
        payload = json.loads(json.dumps(payload_for_prompt, default=str, ensure_ascii=False))
    except Exception:
        payload = dict(payload_for_prompt or {})
    try:
        rag = json.loads(json.dumps(retrieval_context or {}, default=str, ensure_ascii=False))
    except Exception:
        rag = dict(retrieval_context or {})

    def current_tokens():
        p_t, r_t = _estimate_prompt_and_rag_tokens(payload, rag)
        return p_t + r_t

    # Quick no-op
    if current_tokens() <= available_tokens:
        return payload, rag

    # 1) Trim history (oldest first)
    history = payload.get('history') or []
    if isinstance(history, list) and history:
        while history and current_tokens() > available_tokens:
            history.pop(0)
        payload['history'] = history
        if current_tokens() <= available_tokens:
            return payload, rag

    # 2) Reduce retrieved_chunks (RAG secondary) - drop less important chunks first
    retrieved = rag.get('retrieved_chunks') or []
    if isinstance(retrieved, list) and retrieved:
        # keep at least top-1
        while len(retrieved) > 1 and current_tokens() > available_tokens:
            retrieved.pop(-1)
        rag['retrieved_chunks'] = retrieved
        if current_tokens() <= available_tokens:
            return payload, rag

    # 3) Reduce column_samples: drop less helpful columns
    column_samples = payload.get('column_samples') or {}
    if isinstance(column_samples, dict) and column_samples:
        cols = list(column_samples.keys())
        # drop columns from the end until we fit
        while cols and current_tokens() > available_tokens:
            col = cols.pop(-1)
            column_samples.pop(col, None)
        payload['column_samples'] = column_samples
        if current_tokens() <= available_tokens:
            return payload, rag

    # 4) Reduce preview_rows
    preview = payload.get('preview_rows') or []
    if isinstance(preview, list) and preview:
        while preview and current_tokens() > available_tokens:
            preview.pop(-1)
        payload['preview_rows'] = preview
        if current_tokens() <= available_tokens:
            return payload, rag

    # 5) Remove section_fusion and descriptive summaries
    if 'section_fusion' in rag:
        rag.pop('section_fusion', None)
        if current_tokens() <= available_tokens:
            return payload, rag

    # 6) Aggressively truncate chunk summaries in retrieved_chunks
    if isinstance(retrieved, list) and retrieved:
        for i in range(len(retrieved)):
            if current_tokens() <= available_tokens:
                break
            chunk = retrieved[i]
            # keep only minimal fields
            minimal = {
                'chunk_id': chunk.get('chunk_id'),
                'section': chunk.get('section'),
                'row_count': chunk.get('row_count')
            }
            retrieved[i] = minimal
        rag['retrieved_chunks'] = retrieved
        if current_tokens() <= available_tokens:
            return payload, rag

    # If still too big, as a last resort remove rag entirely
    if current_tokens() > available_tokens:
        rag = {'retrieval_policy': 'none', 'retrieved_chunks': []}

    return payload, rag


def _validate_preprocess_llm_output(analysis_result, stage_name, available_columns=None):
    """
    Strict server-side validation for LLM output.
    Returns (is_valid, validation_status, issues).
    """
    available_columns = [str(column) for column in (available_columns or [])]
    issues = []

    if not isinstance(analysis_result, dict):
        issues.append('Resultat LLM non dictionnaire.')
        return False, {
            'chunk_valid': False,
            'schema_valid': False,
            'merge_safe': False,
            'medical_confidence': None,
            'stage': stage_name,
            'issues': issues,
        }, issues

    def _require_type(field_name, expected_type):
        value = analysis_result.get(field_name)
        if not isinstance(value, expected_type):
            issues.append(f'Champ {field_name} invalide ou manquant (type attendu: {expected_type.__name__}).')
            return False
        return True

    def _require_list(field_name):
        return _require_type(field_name, list)

    def _require_dict(field_name):
        return _require_type(field_name, dict)

    def _check_numeric_range(path, value, min_value=None, max_value=None):
        if value is None:
            return True
        if not isinstance(value, (int, float)):
            issues.append(f'Champ numerique invalide: {path}.')
            return False
        if min_value is not None and value < min_value:
            issues.append(f'Champ {path} hors borne minimale {min_value}.')
            return False
        if max_value is not None and value > max_value:
            issues.append(f'Champ {path} hors borne maximale {max_value}.')
            return False
        return True

    schema_ok = True
    schema_ok &= _require_dict('dataset_summary')
    schema_ok &= _require_dict('medical_analysis')
    schema_ok &= _require_dict('missing_values_analysis')
    schema_ok &= _require_dict('outliers_analysis')
    schema_ok &= _require_dict('duplicate_analysis')
    schema_ok &= _require_list('corrections_applied')
    schema_ok &= _require_list('suspect_values')
    schema_ok &= _require_list('remaining_risks')
    schema_ok &= _require_list('recommendations')
    schema_ok &= _require_list('cleaned_dataset_preview')
    schema_ok &= _require_dict('processing_statistics')
    schema_ok &= _require_dict('quality_score')

    if not isinstance(analysis_result.get('summary'), str):
        issues.append('Champ summary invalide ou manquant (string attendu).')
        schema_ok = False

    # Optional but expected for traceability
    if analysis_result.get('correction_plan') is not None and not isinstance(analysis_result.get('correction_plan'), dict):
        issues.append('Champ correction_plan invalide (dict attendu).')
        schema_ok = False

    if not schema_ok:
        return False, {
            'chunk_valid': False,
            'schema_valid': False,
            'merge_safe': False,
            'medical_confidence': None,
            'stage': stage_name,
            'issues': issues,
        }, issues

    dataset_summary = analysis_result.get('dataset_summary') or {}
    processing_statistics = analysis_result.get('processing_statistics') or {}
    quality_score = analysis_result.get('quality_score') or {}
    medical_analysis = analysis_result.get('medical_analysis') or {}
    correction_plan = analysis_result.get('correction_plan') or {}

    business_ok = True
    for key in ['rows', 'row_count', 'duplicate_rows']:
        if key in dataset_summary:
            business_ok &= _check_numeric_range(f'dataset_summary.{key}', dataset_summary.get(key), 0, None)
    for key in ['rows_processed', 'columns_processed', 'chunk_count']:
        if key in processing_statistics:
            business_ok &= _check_numeric_range(f'processing_statistics.{key}', processing_statistics.get(key), 0, None)

    if 'value' in quality_score:
        business_ok &= _check_numeric_range('quality_score.value', quality_score.get('value'), 0, 100)
    if 'confidence' in quality_score:
        business_ok &= _check_numeric_range('quality_score.confidence', quality_score.get('confidence'), 0.0, 1.0)

    if 'confidence' in medical_analysis:
        business_ok &= _check_numeric_range('medical_analysis.confidence', medical_analysis.get('confidence'), 0.0, 1.0)
    if 'medical_confidence' in analysis_result:
        business_ok &= _check_numeric_range('medical_confidence', analysis_result.get('medical_confidence'), 0.0, 1.0)

    merge_safe = True
    if correction_plan:
        required_plan_fields = [
            'rename_columns', 'drop_columns', 'value_mappings', 'fill_missing',
            'type_casts', 'parse_dates', 'trim_whitespace_columns', 'default_values',
        ]
        for field_name in required_plan_fields:
            value = correction_plan.get(field_name)
            expected = dict if field_name in {'rename_columns', 'value_mappings', 'fill_missing', 'type_casts', 'default_values'} else list
            if not isinstance(value, expected):
                issues.append(f'Champ correction_plan.{field_name} invalide (type {expected.__name__} attendu).')
                merge_safe = False

        rename_columns = correction_plan.get('rename_columns') or {}
        drop_columns = correction_plan.get('drop_columns') or []
        if isinstance(rename_columns, dict):
            rename_targets = [str(value) for value in rename_columns.values() if value not in [None, '']]
            if len(rename_targets) != len(set(rename_targets)):
                issues.append('Conflit de renommage: plusieurs colonnes sources ciblent le meme nom.')
                merge_safe = False
            if available_columns:
                for source_name, target_name in rename_columns.items():
                    source_name = str(source_name)
                    target_name = str(target_name)
                    if source_name in drop_columns:
                        issues.append(f'Conflit merge: {source_name} est a la fois renomme et supprime.')
                        merge_safe = False
                    if target_name in available_columns and target_name != source_name:
                        issues.append(f'Conflit merge: cible de renommage deja existante ({target_name}).')
                        merge_safe = False

    if 'limitations' in analysis_result and not isinstance(analysis_result.get('limitations'), list):
        issues.append('Champ limitations invalide (liste attendue).')
        schema_ok = False

    # Derive a normalized medical confidence if not explicitly given.
    medical_confidence = None
    if isinstance(medical_analysis.get('confidence'), (int, float)):
        medical_confidence = float(medical_analysis.get('confidence'))
    elif isinstance(quality_score.get('confidence'), (int, float)):
        medical_confidence = float(quality_score.get('confidence'))
    elif isinstance(quality_score.get('value'), (int, float)):
        medical_confidence = max(0.0, min(1.0, float(quality_score.get('value')) / 100.0))

    if medical_confidence is not None:
        business_ok &= _check_numeric_range('derived_medical_confidence', medical_confidence, 0.0, 1.0)

    valid = schema_ok and business_ok and merge_safe
    validation_status = {
        'chunk_valid': bool(valid),
        'schema_valid': bool(schema_ok),
        'merge_safe': bool(valid and merge_safe),
        'medical_confidence': medical_confidence,
        'stage': stage_name,
        'issues': issues,
        'business_valid': bool(business_ok),
        'available_columns': available_columns[:60],
    }
    return valid, validation_status, issues


def _get_ollama_candidate_bases():
    configured_base = str(os.environ.get('OLLAMA_BASE_URL', '') or os.environ.get('OLLAMA_URL', '')).rstrip('/')
    fallback_base = str(os.environ.get('OLLAMA_FALLBACK_URL', '')).rstrip('/')
    running_in_docker = os.path.exists('/.dockerenv')
    default_base = 'http://ollama:11434' if running_in_docker else 'http://127.0.0.1:11434'

    candidate_bases = []
    for base in [configured_base, fallback_base, default_base]:
        normalized_base = str(base).rstrip('/')
        if normalized_base and normalized_base not in candidate_bases:
            candidate_bases.append(normalized_base)
    return candidate_bases


def _check_ollama_health(timeout_seconds=8):
    endpoint_errors = []
    for ollama_base in _get_ollama_candidate_bases():
        ollama_endpoint = f'{ollama_base}/api/tags'
        req = urllib_request.Request(ollama_endpoint, method='GET')
        try:
            with urllib_request.urlopen(req, timeout=timeout_seconds) as response:
                body = response.read().decode('utf-8')
            payload = json.loads(body)
            models = payload.get('models') if isinstance(payload, dict) else []
            model_names = []
            if isinstance(models, list):
                model_names = [item.get('name') for item in models if isinstance(item, dict) and item.get('name')]
            return {
                'connected': True,
                'base_url': ollama_base,
                'endpoint': ollama_endpoint,
                'models_count': len(model_names),
                'models': model_names[:20],
                'errors': [],
            }
        except urllib_error.HTTPError as error:
            error_body = ''
            try:
                error_body = error.read().decode('utf-8')
            except Exception:
                error_body = ''
            endpoint_errors.append(f'{ollama_endpoint} -> HTTP {error.code}: {error_body or str(error)}')
        except (urllib_error.URLError, TimeoutError, json.JSONDecodeError, ValueError) as error:
            endpoint_errors.append(f'{ollama_endpoint} -> {error}')

    return {
        'connected': False,
        'base_url': None,
        'endpoint': None,
        'models_count': 0,
        'models': [],
        'errors': endpoint_errors,
    }


def _build_deterministic_analysis_fallback(dataframe, technical_profile, reason_message):
    rows_count = int(len(dataframe.index))
    issues = []
    recommendations = []
    suspect_values = []
    cleaned_dataset_preview = _dataframe_to_rows(dataframe.head(5))

    missing_pct = float(technical_profile.get('missing_pct') or 0.0)
    duplicate_rows = int(technical_profile.get('duplicate_rows') or 0)

    if missing_pct > 0:
        issues.append({
            'severity': 'warning',
            'category': 'missing_values',
            'column': '*',
            'rows': max(1, int(round(rows_count * (missing_pct / 100.0)))) if rows_count else 0,
            'explanation': f'Taux de valeurs manquantes detecte: {round(missing_pct, 2)}%.',
        })
        recommendations.append('Completer les valeurs manquantes par strategie deterministe (median/mode).')

    if duplicate_rows > 0:
        issues.append({
            'severity': 'warning',
            'category': 'duplicates',
            'column': '*',
            'rows': duplicate_rows,
            'explanation': f'{duplicate_rows} ligne(s) dupliquee(s) detectee(s).',
        })
        recommendations.append('Verifier et dedoublonner les lignes identiques avant integration.')

    outlier_total = 0
    for numeric_meta in technical_profile.get('numeric_columns_profile', []):
        outlier_total += int(numeric_meta.get('outlier_count') or 0)
    if outlier_total > 0:
        issues.append({
            'severity': 'info',
            'category': 'outliers',
            'column': '*',
            'rows': outlier_total,
            'explanation': f'{outlier_total} valeur(s) numerique(s) potentiellement aberrante(s) detectee(s) (regle IQR).',
        })
        recommendations.append('Examiner les valeurs numeriques extremes avant validation clinique.')

    fill_missing = {}
    trim_columns = []
    parse_dates = []

    for column_meta in technical_profile.get('columns_profile', []):
        column_name = str(column_meta.get('column') or '')
        if not column_name:
            continue
        missing_count = int(column_meta.get('missing_count') or 0)
        dtype_name = str(column_meta.get('dtype') or '').lower()
        if missing_count > 0:
            if 'int' in dtype_name or 'float' in dtype_name:
                fill_missing[column_name] = {'strategy': 'median'}
            else:
                fill_missing[column_name] = {'strategy': 'mode'}
        if dtype_name in {'object', 'string'}:
            trim_columns.append(column_name)
        lowered = column_name.lower()
        if 'date' in lowered or 'naissance' in lowered or 'visite' in lowered:
            parse_dates.append(column_name)

    correction_plan = {
        'fill_missing': fill_missing,
        'parse_dates': parse_dates[:15],
        'type_casts': {},
        'trim_whitespace_columns': trim_columns[:20],
    }

    if not recommendations:
        recommendations.append('Dataset globalement stable selon les controles deterministes locaux.')

    quality_score = max(5, min(100, int(round(100 - (missing_pct * 0.5) - (duplicate_rows * 2) - min(outlier_total, 20)))))

    return {
        'dataset_summary': {
            'rows': int(len(dataframe.index)),
            'columns': int(len(dataframe.columns)),
            'column_names': [str(column) for column in dataframe.columns],
            'missing_cells': int(technical_profile.get('missing_cells') or 0),
            'missing_pct': float(missing_pct),
            'duplicate_rows': int(technical_profile.get('duplicate_rows') or 0),
        },
        'medical_analysis': {
            'summary': 'Analyse locale deterministe sans LLM.',
            'flags': [],
            'notes': [reason_message],
        },
        'missing_values_analysis': {
            'missing_pct': float(missing_pct),
            'strategy': 'deterministic_local',
            'columns': [
                {
                    'column': str(column_meta.get('column') or ''),
                    'missing_count': int(column_meta.get('missing_count') or 0),
                    'missing_pct': float(column_meta.get('missing_pct') or 0.0),
                }
                for column_meta in technical_profile.get('columns_profile', [])[:20]
            ],
        },
        'outliers_analysis': {
            'count': int(outlier_total),
            'method': 'iqr_local',
        },
        'duplicate_analysis': {
            'duplicate_rows': int(duplicate_rows),
            'method': 'exact_duplicate_rows',
        },
        'corrections_applied': [],
        'suspect_values': suspect_values,
        'remaining_risks': list(dict.fromkeys([item['explanation'] for item in issues]))[:5],
        'recommendations': recommendations[:5],
        'cleaned_dataset_preview': cleaned_dataset_preview,
        'processing_statistics': {
            'mode': 'deterministic_fallback',
            'rows_processed': rows_count,
            'columns_processed': int(len(dataframe.columns)),
            'analysis_source': 'pandas',
            'chunks_count': technical_profile.get('chunk_count', 0),
        },
        'quality_score': {
            'value': quality_score,
            'scale': '0-100',
        },
        'quality_score_value': quality_score,
        'summary': 'Analyse realisee via fallback deterministe local (Pandas) suite a indisponibilite LLM.',
        'issues': issues[:6],
        'recommendations': recommendations[:5],
        'correction_plan': correction_plan,
        'corrected_preview_rows': [],
        'column_assessment': [],
        'limitations': [
            'LLM indisponible: fallback deterministe applique.',
            reason_message,
        ],
        'analysis_pack': {
            'pack_type': 'structured_analysis_pack',
            'technical_profile': technical_profile,
            'preview_rows': _dataframe_to_rows(dataframe.head(3)),
            'column_samples': {},
            'meta': {
                'selected_columns_count': int(len(dataframe.columns)),
                'total_columns_count': int(len(dataframe.columns)),
                'preview_rows_count': min(3, int(len(dataframe.index))),
                'column_samples_per_column': 0,
            },
        },
        'model_used': 'deterministic_fallback',
        'attempt': 'fallback_local',
        'stage': 'pass1_diagnostic',
        'second_pass': {'status': 'skipped_fallback'},
        'route': {'mode': 'deterministic', 'label': 'deterministic_only', 'reason': reason_message},
        'pipeline': {
            'stage': 'deterministic_only',
            'chunks_count': technical_profile.get('chunk_count', 0),
            'retrieval': 'skipped',
        },
    }


def _build_preprocess_instruction_block():
    return (
        'Tu es un systeme expert de pretraitement intelligent de donnees medicales specialise en nephrologie, dialyse et analyse clinique. '
        'Tu dois raisonner comme un expert Big Data, un Data Engineer, un Data Scientist medical, un specialiste qualite des donnees et un nephrologue clinique senior. '
        'Analyse l ensemble du dataset, detecte les erreurs, identifie les incoherences medicales, corrige uniquement les anomalies fiables, standardise les donnees et produis une version finale propre. '
        'Tu ne dois jamais halluciner des donnees, inventer des valeurs medicales, modifier silencieusement des donnees critiques ni supprimer automatiquement des donnees importantes. '
        'Si une valeur est ambigue, marque-la comme suspecte, explique pourquoi et reduis le niveau de confiance. '
        'Toutes les sorties doivent etre compactes, deterministes et strictement en JSON valide, sans markdown, sans commentaires et sans texte externe.'
    )


def _call_ollama_qwen_analysis(dataframe, technical_profile, progress_callback=None):
    route = _determine_preprocess_route(technical_profile)
    if route.get('mode') == 'deterministic':
        route = {
            'mode': 'balanced',
            'label': 'llm_only',
            'reason': 'LLM-only mode: no deterministic pandas fallback.',
            'primary_model': os.environ.get('OLLAMA_PREPROCESS_MODEL', os.environ.get('OLLAMA_MODEL', 'qwen2.5:7b-instruct')),
            'fallback_model': os.environ.get('OLLAMA_FALLBACK_MODEL', 'qwen2.5:3b-instruct'),
            'primary_timeout_seconds': _env_int('OLLAMA_PRIMARY_TIMEOUT_SECONDS', min(_env_int('OLLAMA_TIMEOUT_SECONDS', 420), 180)),
            'fallback_timeout_seconds': _env_int('OLLAMA_FALLBACK_TIMEOUT_SECONDS', _env_int('OLLAMA_TIMEOUT_SECONDS', 420)),
            'primary_num_predict': _env_int('OLLAMA_NUM_PREDICT', 32),
            'fallback_num_predict': _env_int('OLLAMA_RETRY_NUM_PREDICT', 24),
            'clinical_complexity_score': 0,
        }
    chunks = _build_preprocess_chunks(dataframe, technical_profile)
    retrieval_context = _build_retrieval_context(
        dataframe,
        chunks,
        technical_profile,
        stage_name='diagnostic',
        max_chunks=int(os.environ.get('RAG_MAX_CHUNKS', '4')),
        progress_callback=progress_callback,
    )
    technical_profile['chunk_count'] = len(chunks)
    technical_profile['chunks'] = retrieval_context.get('chunk_summaries', [])
    technical_profile['retrieved_chunks'] = retrieval_context.get('retrieved_chunks', [])

    model_name = route.get('primary_model') or os.environ.get('OLLAMA_PREPROCESS_MODEL', os.environ.get('OLLAMA_MODEL', 'qwen2.5:7b-instruct'))
    timeout_seconds = int(os.environ.get('OLLAMA_TIMEOUT_SECONDS', '420'))
    primary_timeout_seconds = int(route.get('primary_timeout_seconds') or int(os.environ.get('OLLAMA_PRIMARY_TIMEOUT_SECONDS', str(min(timeout_seconds, 180)))))
    fallback_model = route.get('fallback_model') or os.environ.get('OLLAMA_FALLBACK_MODEL', 'qwen2.5:3b-instruct')
    fallback_timeout_seconds = int(route.get('fallback_timeout_seconds') or int(os.environ.get('OLLAMA_FALLBACK_TIMEOUT_SECONDS', str(timeout_seconds))))
    retry_max_columns = int(os.environ.get('OLLAMA_RETRY_MAX_COLUMNS', '20'))
    retry_preview_rows = int(os.environ.get('OLLAMA_RETRY_PREVIEW_ROWS', '2'))
    retry_num_predict = int(route.get('fallback_num_predict') or os.environ.get('OLLAMA_RETRY_NUM_PREDICT', '24'))
    pass2_num_predict = int(os.environ.get('OLLAMA_PASS2_NUM_PREDICT', str(retry_num_predict)))
    candidate_bases = _get_ollama_candidate_bases()

    prompt_payload = _build_llm_payload(dataframe, technical_profile)
    prompt_payload['rag'] = {
        'chunk_count': len(chunks),
        'retrieved_chunks': retrieval_context.get('retrieved_chunks', []),
        'retrieved_chunks_count': retrieval_context.get('retrieved_chunks_count', 0),
        'section_fusion': retrieval_context.get('section_fusion', []),
        'vector_store': retrieval_context.get('vector_store', {}),
    }
    # Enforce a strict context budget to avoid Ollama truncation/hallucinations.
    try:
        max_context_tokens = int(os.environ.get('MAX_CONTEXT_TOKENS', '8192'))
    except Exception:
        max_context_tokens = 8192
    try:
        system_prompt_tokens = int(os.environ.get('SYSTEM_PROMPT_TOKENS', '1500'))
    except Exception:
        system_prompt_tokens = 1500
    try:
        reserved_output_tokens = int(os.environ.get('CONTEXT_OUTPUT_TOKENS', '1000'))
    except Exception:
        reserved_output_tokens = 1000
    try:
        retry_margin_tokens = int(os.environ.get('CONTEXT_RETRY_MARGIN', '500'))
    except Exception:
        retry_margin_tokens = 500

    available_input_tokens = max(128, max_context_tokens - system_prompt_tokens - reserved_output_tokens - retry_margin_tokens)

    # Estimate tokens and shrink payload/retrieval_context if needed
    before_payload_toks, before_rag_toks = _estimate_prompt_and_rag_tokens(prompt_payload, retrieval_context)
    before_total = before_payload_toks + before_rag_toks
    if before_total > available_input_tokens:
        new_payload, new_rag = _shrink_prompt_for_budget(prompt_payload, retrieval_context, available_input_tokens)
        prompt_payload = new_payload
        prompt_payload['rag'] = new_rag
        # record budget decisions
        technical_profile.setdefault('context_budget', {})
        technical_profile['context_budget'].update({
            'max_context_tokens': max_context_tokens,
            'system_prompt_tokens': system_prompt_tokens,
            'reserved_output_tokens': reserved_output_tokens,
            'retry_margin_tokens': retry_margin_tokens,
            'available_input_tokens': available_input_tokens,
            'before_total_input_tokens': before_total,
            'after_total_input_tokens': _estimate_prompt_and_rag_tokens(prompt_payload, prompt_payload.get('rag'))[0] + _estimate_prompt_and_rag_tokens(prompt_payload, prompt_payload.get('rag'))[1],
            'shrunk': True,
        })
    else:
        technical_profile.setdefault('context_budget', {})
        technical_profile['context_budget'].update({
            'max_context_tokens': max_context_tokens,
            'system_prompt_tokens': system_prompt_tokens,
            'reserved_output_tokens': reserved_output_tokens,
            'retry_margin_tokens': retry_margin_tokens,
            'available_input_tokens': available_input_tokens,
            'before_total_input_tokens': before_total,
            'after_total_input_tokens': before_total,
            'shrunk': False,
        })
    fallback_payload = _shrink_llm_payload(
        prompt_payload,
        max_columns=retry_max_columns,
        max_preview_rows=retry_preview_rows,
        max_samples_per_column=1,
    )

    def _notify_progress(message):
        if callable(progress_callback):
            try:
                progress_callback(message)
            except Exception:
                pass

    def _build_pass1_prompt(payload_for_prompt):
        return (
            _build_preprocess_instruction_block() + ' '
            'Objectif passe 1: produire un diagnostic complet du dataset et une version nettoyee exploitable. '
            'Retourne exclusivement un JSON valide. '
            'Structure obligatoire: {"dataset_summary":{},"medical_analysis":{},"missing_values_analysis":{},"outliers_analysis":{},"duplicate_analysis":{},"corrections_applied":[],"suspect_values":[],"remaining_risks":[],"recommendations":[],"cleaned_dataset_preview":[],"processing_statistics":{},"quality_score":{}}. '
            'Contexte: ' + json.dumps(payload_for_prompt, ensure_ascii=False, default=str)
        )

    def _build_pass2_prompt(payload_for_prompt):
        return (
            _build_preprocess_instruction_block() + ' '
            'Objectif passe 2: proposer un plan de correction deterministe, minimal et traçable. '
            'Retourne exclusivement un JSON valide. '
            'Le JSON doit reprendre la structure de la passe 1 et ajouter un champ top-level correction_plan. '
            'Structure obligatoire: {"dataset_summary":{},"medical_analysis":{},"missing_values_analysis":{},"outliers_analysis":{},"duplicate_analysis":{},"corrections_applied":[],"suspect_values":[],"remaining_risks":[],"recommendations":[],"cleaned_dataset_preview":[],"processing_statistics":{},"quality_score":{},"correction_plan":{"rename_columns":{},"drop_columns":[],"value_mappings":{},"fill_missing":{},"type_casts":{},"parse_dates":[],"trim_whitespace_columns":[],"default_values":{}}}. '
            'Ne jamais inventer de valeurs medicales, ne pas supprimer automatiquement les donnees critiques et ne proposer que des corrections hautement probables. '
            'Contexte: ' + json.dumps(payload_for_prompt, ensure_ascii=False, default=str)
        )

    def _run_stage(stage_name, primary_prompt, fallback_prompt, primary_pack, fallback_pack, primary_predict, fallback_predict):
        attempts = [
            {
                'label': 'primary',
                'model': model_name,
                'timeout_seconds': primary_timeout_seconds,
                'num_predict': primary_predict,
                'analysis_pack': primary_pack,
                'prompt': primary_prompt,
            }
        ]

        if fallback_model and fallback_model != model_name:
            attempts.append(
                {
                    'label': 'fallback',
                    'model': fallback_model,
                    'timeout_seconds': fallback_timeout_seconds,
                    'num_predict': fallback_predict,
                    'analysis_pack': fallback_pack,
                    'prompt': fallback_prompt,
                }
            )

        stage_errors = []
        for attempt in attempts:
            request_payload = {
                'model': attempt['model'],
                'prompt': attempt['prompt'],
                'stream': False,
                'format': 'json',
                'options': {
                    'temperature': 0.1,
                    'num_predict': attempt['num_predict'],
                },
            }
            request_data = json.dumps(request_payload).encode('utf-8')

            for ollama_base in candidate_bases:
                ollama_endpoint = f'{ollama_base}/api/generate'
                req = urllib_request.Request(
                    ollama_endpoint,
                    data=request_data,
                    headers={'Content-Type': 'application/json'},
                    method='POST',
                )

                try:
                    with urllib_request.urlopen(req, timeout=attempt['timeout_seconds']) as response:
                        body = response.read().decode('utf-8')
                    payload = json.loads(body)
                    raw_response = payload.get('response', '{}')
                    parsed_response = _parse_llm_analysis_response(raw_response)
                    # If the model returned invalid JSON, attempt targeted JSON-only retries.
                    json_retry_limit = int(os.environ.get('OLLAMA_JSON_RETRIES', '2'))
                    json_retry_prompt_suffix = (
                        ' REGENERATE ONLY valid JSON now. DO NOT include any explanation or markdown. '
                        'Return strictly the JSON object matching the required structure. '
                        'If you cannot produce valid JSON, return an empty JSON object {}.'
                    )
                    def _is_invalid_json_result(resp):
                        if not isinstance(resp, dict):
                            return True
                        # limitations may include localized strings indicating invalid JSON
                        lim = resp.get('limitations') or []
                        for item in lim:
                            if not isinstance(item, str):
                                continue
                            s = item.lower()
                            # common patterns mentioning JSON invalidity
                            if 'json' in s and ('invalide' in s or 'invalid' in s or 'non valide' in s):
                                return True
                            if 'le modele a repondu' in s and 'json' in s:
                                return True

                        # If parsing failed to populate expected keys, consider it invalid
                        if not resp.get('dataset_summary') and not resp.get('medical_analysis') and not resp.get('summary'):
                            return True
                        return False

                    if _is_invalid_json_result(parsed_response) and json_retry_limit > 0:
                        # use the smaller fallback prompt payload if available
                        regen_base_prompt = attempt.get('prompt')
                        regen_prompt = None
                        try:
                            prev_raw = str(raw_response or '')
                            regen_prompt = (_build_preprocess_instruction_block() + json_retry_prompt_suffix + ' Previous model output: ' + json.dumps({'prev_response': prev_raw}, ensure_ascii=False))
                        except Exception:
                            regen_prompt = (_build_preprocess_instruction_block() + json_retry_prompt_suffix)

                        for retry_idx in range(json_retry_limit):
                            retry_payload = {
                                'model': attempt['model'],
                                'prompt': regen_prompt,
                                'stream': False,
                                'format': 'json',
                                'options': {
                                    'temperature': 0.0,
                                    'num_predict': max(1, int(attempt.get('num_predict', 1) // 2)),
                                },
                            }
                            retry_data = json.dumps(retry_payload).encode('utf-8')
                            try:
                                # Recreate the Request per-retry to ensure the new body is sent
                                retry_req = urllib_request.Request(
                                    ollama_endpoint,
                                    data=retry_data,
                                    headers={'Content-Type': 'application/json'},
                                    method='POST',
                                )
                                with urllib_request.urlopen(retry_req, timeout=attempt['timeout_seconds']) as retry_resp:
                                    retry_body = retry_resp.read().decode('utf-8')
                                retry_payload_json = json.loads(retry_body)
                                retry_raw = retry_payload_json.get('response', '{}')
                                retry_parsed = _parse_llm_analysis_response(retry_raw)
                                if not _is_invalid_json_result(retry_parsed):
                                    parsed_response = retry_parsed
                                    raw_response = retry_raw
                                    break
                            except Exception:
                                continue
                    available_columns = []
                    try:
                        analysis_pack = attempt.get('analysis_pack') or {}
                        if isinstance(analysis_pack, dict) and isinstance(analysis_pack.get('technical_profile'), dict):
                            technical_profile = analysis_pack.get('technical_profile') or {}
                        elif isinstance(analysis_pack, dict) and isinstance(analysis_pack.get('analysis_pack'), dict):
                            technical_profile = (analysis_pack.get('analysis_pack') or {}).get('technical_profile') or {}
                        else:
                            technical_profile = {}
                        available_columns = [
                            str(column.get('name') or column.get('column') or '')
                            for column in (technical_profile.get('columns_profile') or [])
                            if isinstance(column, dict)
                        ]
                        available_columns = [column for column in available_columns if column]
                    except Exception:
                        available_columns = []

                    is_valid, validation_status, validation_issues = _validate_preprocess_llm_output(
                        parsed_response,
                        stage_name=stage_name,
                        available_columns=available_columns,
                    )
                    if not is_valid:
                        stage_errors.append(
                            f"[{stage_name}:{attempt['label']}:{attempt['model']}] validation serveur stricte rejetee: {', '.join(validation_issues[:6])}"
                        )
                        continue

                    parsed_response['validation_status'] = validation_status
                    parsed_response['analysis_pack'] = attempt['analysis_pack']
                    parsed_response['model_used'] = attempt['model']
                    parsed_response['attempt'] = attempt['label']
                    parsed_response['stage'] = stage_name
                    return parsed_response
                except urllib_error.HTTPError as error:
                    error_body = ''
                    try:
                        error_body = error.read().decode('utf-8')
                    except Exception:
                        error_body = ''
                    stage_errors.append(
                        f"[{stage_name}:{attempt['label']}:{attempt['model']}] {ollama_endpoint} -> HTTP {error.code}: {error_body or str(error)}"
                    )
                except (urllib_error.URLError, RemoteDisconnected, ConnectionError, TimeoutError, json.JSONDecodeError, ValueError, OSError) as error:
                    stage_errors.append(
                        f"[{stage_name}:{attempt['label']}:{attempt['model']}] {ollama_endpoint} -> {error}"
                    )

        error_summary = '; '.join(stage_errors) if stage_errors else 'Aucune reponse recue'
        lower_errors = error_summary.lower()
        limitations = []
        if 'timed out' in lower_errors or 'timeout' in lower_errors:
            limitations.append(
                f'Delai depasse: Ollama n a pas repondu dans la fenetre configured (primary={primary_timeout_seconds}s, fallback={fallback_timeout_seconds}s, plafond={timeout_seconds}s).'
            )
        if 'connection refused' in lower_errors or 'errno 111' in lower_errors:
            limitations.append(
                'Connexion refusee: l URL Ollama ciblee n est pas joignable depuis le backend. Verifier OLLAMA_URL et le reseau Docker.'
            )
        if '405' in lower_errors or 'method not allowed' in lower_errors:
            limitations.append('Methode invalide detectee: /api/generate doit etre appele en POST.')
        if not limitations:
            limitations.append('Aucun diagnostic automatique supplementaire disponible.')

        return {
            'unavailable': True,
            'summary': f'Analyse LLM indisponible ({stage_name}): {error_summary}',
            'issues': [],
            'recommendations': [],
            'correction_plan': {},
            'corrected_preview_rows': [],
            'column_assessment': [],
            'limitations': limitations,
            'analysis_pack': primary_pack,
            'model_used': model_name,
            'stage': stage_name,
            'validation_status': {
                'chunk_valid': False,
                'schema_valid': False,
                'merge_safe': False,
                'medical_confidence': None,
                'stage': stage_name,
                'issues': limitations,
            },
        }

    _notify_progress('Analyse LLM - passe 1/2 (diagnostic)...')
    pass1_result = _run_stage(
        stage_name='pass1_diagnostic',
        primary_prompt=_build_pass1_prompt(prompt_payload),
        fallback_prompt=_build_pass1_prompt(fallback_payload),
        primary_pack=prompt_payload,
        fallback_pack=fallback_payload,
        primary_predict=int(os.environ.get('OLLAMA_NUM_PREDICT', '32')),
        fallback_predict=retry_num_predict,
    )

    if pass1_result.get('unavailable'):
        return {
            'unavailable': True,
            'summary': pass1_result.get('summary', 'Passe 1 indisponible.'),
            'issues': [],
            'recommendations': [],
            'correction_plan': {},
            'corrected_preview_rows': [],
            'column_assessment': [],
            'limitations': list(dict.fromkeys((pass1_result.get('limitations') if isinstance(pass1_result.get('limitations'), list) else []) + [
                'Mode LLM-only: aucun fallback Pandas applique.',
            ])),
            'analysis_pack': prompt_payload,
            'model_used': pass1_result.get('model_used'),
            'stage': pass1_result.get('stage', 'pass1_diagnostic'),
            'route': route,
            'section_analyses': retrieval_context.get('section_fusion', []),
            'rag_context': retrieval_context,
            'validation_status': pass1_result.get('validation_status'),
        }

    pass1_issues = pass1_result.get('issues') if isinstance(pass1_result.get('issues'), list) else []
    pass1_recommendations = pass1_result.get('recommendations') if isinstance(pass1_result.get('recommendations'), list) else []
    pass1_limitations = pass1_result.get('limitations') if isinstance(pass1_result.get('limitations'), list) else []

    needs_correction_plan = bool(pass1_result.get('needs_correction_plan')) or bool(pass1_issues)

    if not needs_correction_plan:
        pass1_result['correction_plan'] = pass1_result.get('correction_plan') if isinstance(pass1_result.get('correction_plan'), dict) else {}
        pass1_result['second_pass'] = {'status': 'skipped'}
        pass1_result['route'] = route
        pass1_result['section_analyses'] = retrieval_context.get('section_fusion', [])
        pass1_result['rag_context'] = retrieval_context
        pass1_result['pipeline'] = {
            'stage': 'pass1_diagnostic_completed',
            'chunks_count': len(chunks),
            'retrieved_chunks_count': retrieval_context.get('retrieved_chunks_count', 0),
            'retrieval': retrieval_context.get('retrieval_policy'),
            'vector_store': retrieval_context.get('vector_store', {}),
        }
        return pass1_result

    _notify_progress('Analyse LLM - passe 2/2 (plan de correction)...')
    pass2_primary_payload = {
        'analysis_pack': prompt_payload,
        'diagnostic': {
            'quality_score': pass1_result.get('quality_score'),
            'summary': pass1_result.get('summary'),
            'issues': pass1_issues[:6],
            'recommendations': pass1_recommendations[:5],
        },
    }
    pass2_fallback_payload = {
        'analysis_pack': fallback_payload,
        'diagnostic': pass2_primary_payload['diagnostic'],
    }

    pass2_result = _run_stage(
        stage_name='pass2_correction_plan',
        primary_prompt=_build_pass2_prompt(pass2_primary_payload),
        fallback_prompt=_build_pass2_prompt(pass2_fallback_payload),
        primary_pack=pass2_primary_payload,
        fallback_pack=pass2_fallback_payload,
        primary_predict=pass2_num_predict,
        fallback_predict=min(pass2_num_predict, retry_num_predict),
    )

    final_result = {
        'quality_score': pass1_result.get('quality_score'),
        'summary': pass1_result.get('summary'),
        'issues': pass1_issues,
        'recommendations': pass1_recommendations,
        'correction_plan': {},
        'corrected_preview_rows': pass1_result.get('corrected_preview_rows', []),
        'column_assessment': pass1_result.get('column_assessment', []),
        'limitations': pass1_limitations,
        'analysis_pack': pass1_result.get('analysis_pack', prompt_payload),
        'model_used': pass1_result.get('model_used'),
        'attempt': pass1_result.get('attempt'),
        'second_pass': {},
        'route': route,
        'section_analyses': retrieval_context.get('section_fusion', []),
        'rag_context': retrieval_context,
        'pipeline': {
            'stage': 'pass2_correction_plan',
            'chunks_count': len(chunks),
            'retrieved_chunks_count': retrieval_context.get('retrieved_chunks_count', 0),
            'retrieval': retrieval_context.get('retrieval_policy'),
            'vector_store': retrieval_context.get('vector_store', {}),
        },
    }

    final_result['validation_status'] = pass1_result.get('validation_status') or {
        'chunk_valid': False,
        'schema_valid': False,
        'merge_safe': False,
        'medical_confidence': None,
        'stage': 'pass1_diagnostic',
        'issues': ['Validation manquante sur la passe 1.'],
    }

    if pass2_result.get('unavailable'):
        final_result['limitations'] = list(dict.fromkeys(final_result['limitations'] + [
            'Passe 2 indisponible: plan de correction vide utilise.'
        ] + (pass2_result.get('limitations') if isinstance(pass2_result.get('limitations'), list) else [])))
        final_result['second_pass'] = {
            'status': 'failed',
            'model_used': pass2_result.get('model_used'),
            'attempt': pass2_result.get('attempt'),
        }
        final_result['validation_status_pass2'] = pass2_result.get('validation_status')
        return final_result

    correction_plan = pass2_result.get('correction_plan') if isinstance(pass2_result.get('correction_plan'), dict) else {}
    final_result['correction_plan'] = correction_plan
    final_result['second_pass'] = {
        'status': 'completed',
        'model_used': pass2_result.get('model_used'),
        'attempt': pass2_result.get('attempt'),
    }
    final_result['validation_status'] = pass2_result.get('validation_status') or final_result.get('validation_status')
    final_result['validation_status_pass2'] = pass2_result.get('validation_status')
    if isinstance(pass2_result.get('limitations'), list) and pass2_result.get('limitations'):
        final_result['limitations'] = list(dict.fromkeys(final_result['limitations'] + pass2_result.get('limitations')))

    return final_result


def _resolve_llm_column_name(column_name, rename_map, available_columns):
    if column_name in available_columns:
        return column_name
    if column_name in rename_map:
        renamed = rename_map[column_name]
        if renamed in available_columns:
            return renamed
    normalized_target = normalize_header(column_name)
    for candidate in available_columns:
        if normalize_header(candidate) == normalized_target:
            return candidate
    return None


def _apply_llm_fill_strategy(series, strategy_spec):
    if strategy_spec in [None, '']:
        return series
    if not isinstance(strategy_spec, dict):
        return series.fillna(strategy_spec)

    strategy = str(strategy_spec.get('strategy', '')).lower()
    value = strategy_spec.get('value')

    if strategy == 'constant':
        return series.fillna(value)
    if strategy == 'mode':
        mode_values = series.mode(dropna=True)
        if not mode_values.empty:
            return series.fillna(mode_values.iloc[0])
        return series
    if strategy == 'mean':
        numeric = pd.to_numeric(series, errors='coerce')
        if numeric.notna().any():
            return series.fillna(float(numeric.mean()))
        return series
    if strategy == 'median':
        numeric = pd.to_numeric(series, errors='coerce')
        if numeric.notna().any():
            return series.fillna(float(numeric.median()))
        return series
    if strategy == 'forward_fill':
        return series.fillna(method='ffill')
    if strategy == 'backward_fill':
        return series.fillna(method='bfill')

    return series.fillna(value)


def _apply_llm_correction_plan(dataframe, llm_analysis):
    if not isinstance(llm_analysis, dict):
        return dataframe.copy(), []

    correction_plan = llm_analysis.get('correction_plan') or {}
    if not isinstance(correction_plan, dict):
        return dataframe.copy(), []

    corrected = dataframe.copy()
    applied_actions = []

    rename_columns = correction_plan.get('rename_columns') or {}
    rename_map = {}
    if isinstance(rename_columns, dict):
        rename_map = {str(source): str(target) for source, target in rename_columns.items() if source and target}
        if rename_map:
            existing_renames = {source: target for source, target in rename_map.items() if source in corrected.columns}
            if existing_renames:
                corrected = corrected.rename(columns=existing_renames)
                applied_actions.append({'action': 'rename_columns', 'count': len(existing_renames)})

    available_columns = list(corrected.columns)

    drop_columns = correction_plan.get('drop_columns') or []
    if isinstance(drop_columns, list):
        columns_to_drop = [column for column in drop_columns if column in available_columns]
        if columns_to_drop:
            corrected = corrected.drop(columns=columns_to_drop)
            applied_actions.append({'action': 'drop_columns', 'count': len(columns_to_drop)})
            available_columns = list(corrected.columns)

    trim_columns = correction_plan.get('trim_whitespace_columns') or []
    if isinstance(trim_columns, list):
        trimmed_count = 0
        for column_name in trim_columns:
            resolved = _resolve_llm_column_name(str(column_name), rename_map, available_columns)
            if not resolved or corrected[resolved].dtype != object:
                continue
            corrected[resolved] = corrected[resolved].apply(lambda value: value.strip() if isinstance(value, str) else value)
            trimmed_count += 1
        if trimmed_count:
            applied_actions.append({'action': 'trim_whitespace', 'count': trimmed_count})

    value_mappings = correction_plan.get('value_mappings') or {}
    if isinstance(value_mappings, dict):
        mapping_count = 0
        for column_name, mapping in value_mappings.items():
            resolved = _resolve_llm_column_name(str(column_name), rename_map, available_columns)
            if not resolved or not isinstance(mapping, dict):
                continue
            corrected[resolved] = corrected[resolved].replace(mapping)
            mapping_count += 1
        if mapping_count:
            applied_actions.append({'action': 'value_mappings', 'count': mapping_count})

    type_casts = correction_plan.get('type_casts') or {}
    if isinstance(type_casts, dict):
        cast_count = 0
        for column_name, target_type in type_casts.items():
            resolved = _resolve_llm_column_name(str(column_name), rename_map, available_columns)
            if not resolved:
                continue
            target = str(target_type).lower()
            if target in {'numeric', 'number', 'float', 'decimal'}:
                corrected[resolved] = pd.to_numeric(corrected[resolved], errors='coerce')
            elif target in {'integer', 'int'}:
                numeric = pd.to_numeric(corrected[resolved], errors='coerce')
                corrected[resolved] = numeric.round().astype('Int64')
            elif target in {'date', 'datetime'}:
                parsed = pd.to_datetime(corrected[resolved], errors='coerce', dayfirst=True)
                corrected[resolved] = parsed.dt.date
            elif target in {'string', 'text'}:
                corrected[resolved] = corrected[resolved].astype('string')
            cast_count += 1
        if cast_count:
            applied_actions.append({'action': 'type_casts', 'count': cast_count})

    parse_dates = correction_plan.get('parse_dates') or []
    if isinstance(parse_dates, list):
        parsed_count = 0
        for column_name in parse_dates:
            resolved = _resolve_llm_column_name(str(column_name), rename_map, available_columns)
            if not resolved:
                continue
            corrected[resolved] = pd.to_datetime(corrected[resolved], errors='coerce', dayfirst=True).dt.date
            parsed_count += 1
        if parsed_count:
            applied_actions.append({'action': 'parse_dates', 'count': parsed_count})

    fill_missing = correction_plan.get('fill_missing') or {}
    if isinstance(fill_missing, dict):
        filled_count = 0
        for column_name, strategy_spec in fill_missing.items():
            resolved = _resolve_llm_column_name(str(column_name), rename_map, available_columns)
            if not resolved:
                continue
            corrected[resolved] = _apply_llm_fill_strategy(corrected[resolved], strategy_spec)
            filled_count += 1
        if filled_count:
            applied_actions.append({'action': 'fill_missing', 'count': filled_count})

    default_values = correction_plan.get('default_values') or {}
    if isinstance(default_values, dict):
        default_count = 0
        for column_name, default_value in default_values.items():
            resolved = _resolve_llm_column_name(str(column_name), rename_map, available_columns)
            if not resolved:
                continue
            corrected[resolved] = corrected[resolved].fillna(default_value)
            default_count += 1
        if default_count:
            applied_actions.append({'action': 'default_values', 'count': default_count})

    return corrected, applied_actions


def _build_preprocess_report(dataframe, technical_profile, llm_analysis=None, corrected_df=None, applied_actions=None):
    llm_analysis = llm_analysis or {}
    corrected_df = corrected_df if isinstance(corrected_df, pd.DataFrame) else dataframe
    applied_actions = applied_actions or []

    issues = llm_analysis.get('issues') if isinstance(llm_analysis, dict) else []
    if not isinstance(issues, list):
        issues = []

    severity_count = {'critical': 0, 'warning': 0, 'info': 0}
    for item in issues:
        severity = str(item.get('severity', 'info')).lower()
        severity_count[severity] = severity_count.get(severity, 0) + 1

    quality_score = llm_analysis.get('quality_score') if isinstance(llm_analysis, dict) else None
    if not isinstance(quality_score, (int, float)):
        raw_score = 100 - (severity_count.get('critical', 0) * 15) - (severity_count.get('warning', 0) * 6) - (severity_count.get('info', 0) * 2)
        quality_score = max(5, min(100, int(raw_score)))
    else:
        quality_score = max(0, min(100, int(round(float(quality_score)))))

    recommendations = []
    for recommendation in (llm_analysis.get('recommendations') or []):
        recommendation_text = str(recommendation)
        if recommendation_text not in recommendations:
            recommendations.append(recommendation_text)

    if not recommendations:
        recommendations.append('Aucune recommandation fournie par le modele.')

    corrected_preview_rows = _dataframe_to_rows(corrected_df.head(20))

    return {
        'summary': {
            'rows': int(len(dataframe.index)),
            'columns': int(len(dataframe.columns)),
            'quality_score': quality_score,
            'total_issues': len(issues),
            'severity_count': severity_count,
            'corrected_rows': int(len(corrected_df.index)),
            'applied_corrections_count': len(applied_actions),
        },
        'dataset_profile': technical_profile,
        'issues': issues,
        'recommendations': recommendations,
        'correction_plan': llm_analysis.get('correction_plan') or {},
        'applied_corrections': applied_actions,
        'llm_analysis': llm_analysis,
        'corrected_preview_rows': corrected_preview_rows,
        'llm_preview_rows': llm_analysis.get('corrected_preview_rows') if isinstance(llm_analysis, dict) else [],
        'analysis_pack': llm_analysis.get('analysis_pack') if isinstance(llm_analysis, dict) else None,
    }


def _integrate_dataframe_into_patients(dataframe, request_user=None, source_file_name='preprocessed_dataset'):
    headers = list(dataframe.columns)
    template = upsert_template_from_headers(headers, None, source_file_name, create_fields=False)

    created_count = 0
    row_errors = []
    auto_increment_state = initialize_auto_increment_state()
    all_dynamic_columns = {}
    existing_template_keys = set(template.fields.values_list('key', flat=True)) if template else set()

    for header in headers:
        header_str = str(header).strip()
        if not header_str:
            continue
        normalized = normalize_header(header_str)
        if not normalized or normalized in existing_template_keys:
            continue
        if normalized in _COLUMNS_FUSED_INTO_COMORBIDITE_LISTE or normalized in COLUMN_MAPPING:
            continue
        all_dynamic_columns.setdefault(normalized, header_str)

    for index, row in dataframe.iterrows():
        payload = build_patient_payload(row)
        detected = payload.pop('_dynamic_columns_detected', set())
        for norm_key in detected:
            if norm_key not in all_dynamic_columns:
                original_name = next(
                    (str(h).strip() for h in headers if normalize_header(str(h).strip()) == norm_key),
                    norm_key,
                )
                all_dynamic_columns[norm_key] = original_name

        payload = apply_automatic_schema_fields(
            payload,
            auto_increment_state=auto_increment_state,
            current_user=request_user,
        )
        payload = ensure_required_identity_fields(payload)
        payload = ensure_incremental_identifiers(
            payload,
            auto_increment_state=auto_increment_state,
            force_generated=True,
        )

        serializer = PatientSerializer(data=payload)
        if serializer.is_valid():
            serializer.save()
            created_count += 1
        else:
            row_errors.append({'row': index + 2, 'errors': serializer.errors})

    dynamic_columns_info = []
    if all_dynamic_columns and template:
        existing_keys = set(template.fields.values_list('key', flat=True))
        new_fields = []
        for norm_key, original_name in all_dynamic_columns.items():
            if norm_key in existing_keys:
                template.fields.filter(key=norm_key).update(source_hint='dynamic_column')
            else:
                new_fields.append(
                    PatientFormField(
                        template=template,
                        key=norm_key,
                        label=original_name,
                        field_type='text_short',
                        order=10000 + len(new_fields),
                        choices=[],
                        source_hint='dynamic_column',
                        is_required=False,
                    )
                )
            dynamic_columns_info.append({
                'key': norm_key,
                'label': original_name,
                'is_new': norm_key not in existing_keys,
            })

        if new_fields:
            PatientFormField.objects.bulk_create(new_fields)

    if created_count > 0 or all_dynamic_columns:
        try:
            refresh_postgres_flat_view(template)
        except Exception:
            pass

    template_data = PatientFormTemplateSerializer(template).data
    status_code = status.HTTP_201_CREATED if created_count else status.HTTP_200_OK
    return {
        'mode': 'data',
        'template': template_data,
        'fields_created': len(template_data.get('fields', [])),
        'patients_created': created_count,
        'errors': row_errors,
        'dynamic_columns': dynamic_columns_info,
        'dynamic_columns_count': len(dynamic_columns_info),
        'new_dynamic_columns_count': sum(1 for c in dynamic_columns_info if c.get('is_new')),
    }, status_code


# ============================================================
# VUES API
# ============================================================

class PatientListCreateView(APIView):
    permission_classes = [CanViewPatients]

    def get_queryset(self, request):
        queryset = Patient.objects.all()
        search = request.query_params.get('search', '').strip()
        id_patient = request.query_params.get('id_patient', '').strip()
        sexe = request.query_params.get('sexe', '').strip()
        age_min = request.query_params.get('age_min', '').strip()
        age_max = request.query_params.get('age_max', '').strip()
        date_naissance = request.query_params.get('date_naissance', '').strip()
        statut_inclusion = request.query_params.get('statut_inclusion', '').strip()
        infection = request.query_params.get('infection', '').strip().lower()
        hemorrhage = request.query_params.get('hemorrhage', '').strip().lower()
        avf_created = request.query_params.get('avf_created', '').strip().lower()

        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) | Q(prenom__icontains=search) |
                Q(id_patient__icontains=search) | Q(maladie__icontains=search) |
                Q(telephone__icontains=search) | Q(adresse__icontains=search)
            )
        if id_patient:
            queryset = queryset.filter(id_patient__icontains=id_patient)
        if sexe:
            queryset = queryset.filter(sexe__iexact=sexe)
        if age_min:
            queryset = queryset.filter(age__gte=age_min)
        if age_max:
            queryset = queryset.filter(age__lte=age_max)
        if date_naissance:
            queryset = queryset.filter(date_naissance=date_naissance)
        if statut_inclusion:
            queryset = queryset.filter(statut_inclusion__icontains=statut_inclusion)

        return queryset

    def get(self, request):
        serializer = PatientSerializer(self.get_queryset(request), many=True)
        return Response(serializer.data)

    def post(self, request):
        before_data = None
        payload = request.data.copy()
        payload = apply_automatic_schema_fields(payload, current_user=request.user)
        payload = ensure_incremental_identifiers(payload)
        serializer = PatientSerializer(data=payload)
        if serializer.is_valid():
            patient = serializer.save()
            AuditLog.objects.create(
                utilisateur=request.user if request.user.is_authenticated else None,
                action=f"CREATION_PATIENT: patient {patient.id_patient or patient.id} cree",
                entite='Patient',
                entite_id=patient.id,
                adresse_ip=request.META.get('REMOTE_ADDR'),
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PatientExportExcelView(APIView):
    permission_classes = [CanViewPatients]

    def get(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'patients'

        headers = [
            'id_patient', 'nom', 'prenom', 'age', 'sexe',
            'demographie_sexe', 'demographie_age_ans', 'demographie_date_naissance',
            'irc_etiologie_principale', 'dialyse_modalite_initiale', 'dialyse_date_debut',
            'devenir_statut', 'devenir_date_deces', 'devenir_cause_deces',
        ]
        worksheet.append(headers)

        for patient in Patient.objects.all():
            row = [
                patient.id_patient, patient.nom, patient.prenom, patient.age, patient.sexe,
                patient.demographie_sexe, patient.demographie_age_ans,
                patient.demographie_date_naissance,
                patient.irc_etiologie_principale, patient.dialyse_modalite_initiale,
                patient.dialyse_date_debut,
                patient.devenir_statut, patient.devenir_date_deces, patient.devenir_cause_deces,
            ]
            worksheet.append(row)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="patients_export.xlsx"'
        workbook.save(response)
        return response


class PatientDetailView(APIView):
    permission_classes = [CanViewPatients]

    def get_object(self, pk):
        return get_object_or_404(Patient, pk=pk)

    def get(self, request, pk):
        serializer = PatientSerializer(self.get_object(pk))
        return Response(serializer.data)

    def put(self, request, pk):
        patient = self.get_object(pk)
        before_data = PatientSerializer(patient).data
        serializer = PatientSerializer(patient, data=request.data, partial=True)
        if serializer.is_valid():
            updated_patient = serializer.save()
            after_data = PatientSerializer(updated_patient).data
            AuditLog.objects.create(
                utilisateur=request.user if request.user.is_authenticated else None,
                action=f"MODIFICATION_PATIENT: {_describe_patient_changes(before_data, after_data)}",
                entite='Patient',
                entite_id=updated_patient.id,
                adresse_ip=request.META.get('REMOTE_ADDR'),
            )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        patient = self.get_object(pk)
        before_data = PatientSerializer(patient).data
        serializer = PatientSerializer(patient, data=request.data, partial=True)
        if serializer.is_valid():
            updated_patient = serializer.save()
            after_data = PatientSerializer(updated_patient).data
            AuditLog.objects.create(
                utilisateur=request.user if request.user.is_authenticated else None,
                action=f"MODIFICATION_PATIENT: {_describe_patient_changes(before_data, after_data)}",
                entite='Patient',
                entite_id=updated_patient.id,
                adresse_ip=request.META.get('REMOTE_ADDR'),
            )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        patient = self.get_object(pk)
        AuditLog.objects.create(
            utilisateur=request.user if request.user.is_authenticated else None,
            action=f"SUPPRESSION_PATIENT: patient {patient.id_patient or patient.id} supprime",
            entite='Patient',
            entite_id=patient.id,
            adresse_ip=request.META.get('REMOTE_ADDR'),
        )
        patient.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class PatientBulkPurgeView(APIView):
    permission_classes = [IsAdminOrChefService]

    def delete(self, request):
        deleted_count, _ = Patient.objects.all().delete()
        return Response({'deleted_count': deleted_count}, status=status.HTTP_200_OK)


class PatientDynamicColumnsCleanupView(APIView):
    permission_classes = [IsAdminOrChefService]

    def post(self, request):
        template = get_active_template()
        if not template:
            return Response(
                {
                    'template_found': False,
                    'removed_count': 0,
                    'removed_keys': [],
                },
                status=status.HTTP_200_OK,
            )

        dynamic_fields = template.fields.filter(
            source_hint__in=['dynamic_column', 'auto_detected_from_data_import']
        )
        if not dynamic_fields.exists():
            return Response(
                {
                    'template_found': True,
                    'removed_count': 0,
                    'removed_keys': [],
                },
                status=status.HTTP_200_OK,
            )

        used_extra_keys = set()
        for extra_payload in Patient.objects.values_list('extra_data', flat=True):
            if isinstance(extra_payload, dict):
                used_extra_keys.update(extra_payload.keys())

        stale_fields_qs = dynamic_fields.exclude(key__in=used_extra_keys)
        removed_keys = list(stale_fields_qs.values_list('key', flat=True))
        removed_count = len(removed_keys)
        if removed_count > 0:
            stale_fields_qs.delete()
            try:
                refresh_postgres_flat_view(template)
            except Exception:
                pass

        return Response(
            {
                'template_found': True,
                'removed_count': removed_count,
                'removed_keys': removed_keys,
            },
            status=status.HTTP_200_OK,
        )


class PatientPreprocessAnalyzeView(APIView):
    permission_classes = [CanViewPatients]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        from patients.tasks import analyze_preprocess_async
        import tempfile

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'Fichier CSV/Excel requis.'}, status=status.HTTP_400_BAD_REQUEST)

        # Créer un ID unique pour cette session
        session_id = uuid.uuid4().hex
        source_file_name = uploaded_file.name

        # Sauvegarder le fichier temporairement
        try:
            temp_dir = '/tmp/preprocess'
            os.makedirs(temp_dir, exist_ok=True)
            temp_file_path = os.path.join(temp_dir, f'{session_id}_{source_file_name}')
            with open(temp_file_path, 'wb') as temp_file:
                for chunk in uploaded_file.chunks():
                    temp_file.write(chunk)
        except Exception as error:
            return Response({'error': f'Erreur lors de la sauvegarde du fichier: {error}'}, status=status.HTTP_400_BAD_REQUEST)

        # Initialiser la session avec statut "pending"
        session = {
            'id': session_id,
            'created_at': timezone.now().isoformat(),
            'created_by': resolve_entry_user_label(request.user),
            'source_file_name': source_file_name,
            'columns': [],
            'original_rows': [],
            'corrected_rows': [],
            'report': {},
            'change_log': [],
            'status': 'pending',
            'error': None,
            'progress_message': 'Initialisation du traitement...',
        }
        _save_preprocess_session(session)

        # Dispatcher la task Celery de manière asynchrone
        try:
            analyze_preprocess_async.delay(
                session_id=session_id,
                file_path=temp_file_path,
                user_id=request.user.id,
                use_llm=str(request.data.get('use_llm', 'true')).lower() not in ['0', 'false', 'no', 'non']
            )
        except Exception as error:
            print(f"Erreur lors du lancement de la task Celery: {error}")
            session['status'] = 'error'
            session['error'] = f'Erreur Celery: {error}'
            _save_preprocess_session(session)
            return Response({'error': f'Erreur lors du lancement du traitement: {error}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(
            {
                'preprocess_id': session_id,
                'status': 'pending',
                'message': 'Analyse en cours... Veuillez vérifier le statut.',
                'source_file_name': source_file_name,
            },
            status=status.HTTP_202_ACCEPTED,
        )


class PatientPreprocessStatusView(APIView):
    """Track the status of an ongoing preprocessing job."""
    permission_classes = [CanViewPatients]

    def get(self, request, preprocess_id):
        try:
            session = _load_preprocess_session(preprocess_id)
        except Exception:
            return Response({'error': 'Session non trouvée.'}, status=status.HTTP_404_NOT_FOUND)

        if session.get('status') == 'completed':
            report = session.get('report', {})
            return Response(
                {
                    'preprocess_id': preprocess_id,
                    'status': 'completed',
                    'report': report,
                    'columns': session.get('columns', []),
                    'row_count': len(session.get('corrected_rows', [])),
                    'original_preview_rows': session.get('original_rows', [])[:20],
                    'preview_rows': session.get('corrected_rows', [])[:20],
                    'corrected_preview_rows': report.get('corrected_preview_rows', [])[:20],
                    'dataset_profile': {'columns': len(session.get('columns', [])), 'rows': len(session.get('corrected_rows', []))},
                    'source_file_name': session.get('source_file_name'),
                },
                status=status.HTTP_200_OK,
            )
        elif session.get('status') == 'error':
            return Response(
                {
                    'preprocess_id': preprocess_id,
                    'status': 'error',
                    'error': session.get('error', 'Erreur inconnue'),
                    'message': f"Erreur lors de l'analyse: {session.get('error')}",
                },
                status=status.HTTP_200_OK,
            )
        else:  # pending or in_progress
            return Response(
                {
                    'preprocess_id': preprocess_id,
                    'status': session.get('status', 'pending'),
                    'progress_message': session.get('progress_message', 'Traitement en cours...'),
                    'message': 'Analyse en cours, veuillez patienter...',
                },
                status=status.HTTP_200_OK,
            )


class PatientPreprocessHealthView(APIView):
    # Allow unauthenticated access so the UI can display Ollama status before login
    permission_classes = [AllowAny]

    def get(self, request):
        timeout_seconds = int(os.environ.get('OLLAMA_HEALTH_TIMEOUT_SECONDS', '8'))
        health = _check_ollama_health(timeout_seconds=timeout_seconds)
        model_name = os.environ.get('OLLAMA_MODEL', 'qwen2.5:7b-instruct')
        if health.get('connected'):
            return Response(
                {
                    'connected': True,
                    'configured_model': model_name,
                    'base_url': health.get('base_url'),
                    'endpoint': health.get('endpoint'),
                    'models_count': health.get('models_count', 0),
                    'models': health.get('models', []),
                    'message': 'Ollama connecté.',
                },
                status=status.HTTP_200_OK,
            )

        errors = health.get('errors') or []
        return Response(
            {
                'connected': False,
                'configured_model': model_name,
                'base_url': None,
                'endpoint': None,
                'models_count': 0,
                'models': [],
                'message': 'Ollama indisponible depuis le backend.',
                'errors': errors,
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )


class PatientPreprocessSessionView(APIView):
    permission_classes = [CanViewPatients]

    def get(self, request, session_id):
        session = _load_preprocess_session(session_id)
        if not session:
            return Response({'error': 'Session de pretraitement introuvable.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(
            {
                'preprocess_id': session.get('id'),
                'source_file_name': session.get('source_file_name'),
                'report': session.get('report', {}),
                'columns': session.get('columns', []),
                'row_count': len(session.get('corrected_rows', [])),
                'preview_rows': session.get('corrected_rows', [])[:50],
                'change_log': session.get('change_log', [])[-20:],
            },
            status=status.HTTP_200_OK,
        )


class PatientPreprocessRowsView(APIView):
    permission_classes = [CanViewPatients]

    def post(self, request, session_id):
        session = _load_preprocess_session(session_id)
        if not session:
            return Response({'error': 'Session de pretraitement introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        row_payload = request.data.get('row') or {}
        columns = session.get('columns', [])
        row = {column: row_payload.get(column) for column in columns}

        for key, value in row_payload.items():
            if key not in row:
                row[key] = value
                if key not in columns:
                    columns.append(key)

        session.setdefault('corrected_rows', []).append(row)
        session['columns'] = columns
        session.setdefault('change_log', []).append(
            {
                'timestamp': timezone.now().isoformat(),
                'user': resolve_entry_user_label(request.user),
                'action': 'create_row',
                'row_index': len(session['corrected_rows']) - 1,
            }
        )
        _save_preprocess_session(session)
        return Response({'row_count': len(session['corrected_rows'])}, status=status.HTTP_200_OK)


class PatientPreprocessRowDetailView(APIView):
    permission_classes = [CanViewPatients]

    def patch(self, request, session_id, row_index):
        session = _load_preprocess_session(session_id)
        if not session:
            return Response({'error': 'Session de pretraitement introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        rows = session.get('corrected_rows', [])
        if row_index < 0 or row_index >= len(rows):
            return Response({'error': 'Index de ligne invalide.'}, status=status.HTTP_400_BAD_REQUEST)

        row_payload = request.data.get('row') or {}
        current_row = rows[row_index]
        for key, value in row_payload.items():
            current_row[key] = value
            if key not in session.get('columns', []):
                session['columns'].append(key)

        session.setdefault('change_log', []).append(
            {
                'timestamp': timezone.now().isoformat(),
                'user': resolve_entry_user_label(request.user),
                'action': 'update_row',
                'row_index': row_index,
            }
        )
        _save_preprocess_session(session)
        return Response({'row': current_row}, status=status.HTTP_200_OK)

    def delete(self, request, session_id, row_index):
        session = _load_preprocess_session(session_id)
        if not session:
            return Response({'error': 'Session de pretraitement introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        rows = session.get('corrected_rows', [])
        if row_index < 0 or row_index >= len(rows):
            return Response({'error': 'Index de ligne invalide.'}, status=status.HTTP_400_BAD_REQUEST)

        rows.pop(row_index)
        session.setdefault('change_log', []).append(
            {
                'timestamp': timezone.now().isoformat(),
                'user': resolve_entry_user_label(request.user),
                'action': 'delete_row',
                'row_index': row_index,
            }
        )
        _save_preprocess_session(session)
        return Response({'row_count': len(rows)}, status=status.HTTP_200_OK)


class PatientPreprocessExportView(APIView):
    permission_classes = [CanViewPatients]

    def get(self, request, session_id):
        session = _load_preprocess_session(session_id)
        if not session:
            return Response({'error': 'Session de pretraitement introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        source = str(request.query_params.get('source', 'corrected')).lower()
        rows_key = 'original_rows' if source == 'original' else 'corrected_rows'
        rows = session.get(rows_key, [])
        columns = session.get('columns', [])
        dataframe = _rows_to_dataframe(rows, columns=columns)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=False, sheet_name='preprocess_dataset')
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="preprocess_{session_id}_{source}.xlsx"'
        return response


class PatientPreprocessIntegrateView(APIView):
    permission_classes = [CanViewPatients]

    def post(self, request, session_id):
        session = _load_preprocess_session(session_id)
        if not session:
            return Response({'error': 'Session de pretraitement introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        source = str(request.data.get('source', 'corrected')).lower()
        rows_key = 'original_rows' if source == 'original' else 'corrected_rows'
        rows = session.get(rows_key, [])
        columns = session.get('columns', [])
        dataframe = _rows_to_dataframe(rows, columns=columns)

        result_payload, result_status = _integrate_dataframe_into_patients(
            dataframe,
            request_user=request.user,
            source_file_name=session.get('source_file_name') or f'preprocess_{session_id}.xlsx',
        )
        result_payload['preprocess_id'] = session_id
        result_payload['integration_source'] = source
        result_payload['quality_score'] = session.get('report', {}).get('summary', {}).get('quality_score')
        return Response(result_payload, status=result_status)


class PatientImportExcelView(APIView):
    permission_classes = [CanViewPatients]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'error': 'Fichier Excel requis'}, status=status.HTTP_400_BAD_REQUEST)

        source_file_name = getattr(excel_file, 'name', '')
        lower_name = str(source_file_name).lower()
        worksheet = None
        dataframe = None

        if lower_name.endswith('.xls'):
            try:
                excel_file.seek(0)
                dataframe = pd.read_excel(excel_file, engine='xlrd')
            except Exception as error:
                return Response(
                    {'error': f'Impossible de lire le fichier Excel (.xls): {error}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            try:
                excel_file.seek(0)
                workbook = load_workbook(excel_file, data_only=False)
            except Exception as error:
                return Response(
                    {'error': f'Impossible de lire le fichier Excel: {error}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            worksheet = workbook[workbook.sheetnames[0]]

            if is_schema_template_sheet(worksheet):
                template, fields_created = parse_schema_from_template_sheet(worksheet, source_file_name)
                template_data = PatientFormTemplateSerializer(template).data
                return Response(
                    {
                        'mode': 'schema',
                        'template': template_data,
                        'fields_created': fields_created,
                        'patients_created': 0,
                        'errors': [],
                    },
                    status=status.HTTP_201_CREATED,
                )

            try:
                excel_file.seek(0)
                # parse_dates=True aide pandas à reconnaître les colonnes de dates
                dataframe = pd.read_excel(excel_file, parse_dates=True)
            except Exception as error:
                return Response(
                    {'error': f'Impossible de lire les donnees du fichier Excel: {error}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if dataframe is not None:
            dataframe = dataframe.loc[
                :, ~dataframe.columns.astype(str).str.match(r'^(Unnamed|unnamed)(:.*)?$')
            ]

        headers = list(dataframe.columns)
        template = upsert_template_from_headers(headers, worksheet, source_file_name, create_fields=False)

        created_count = 0
        row_errors = []
        auto_increment_state = initialize_auto_increment_state()
        # Accumuler toutes les colonnes dynamiques rencontrées sur l'ensemble des lignes
        # Pré-détecter les colonnes non-mappées/non-fixes à partir des headers
        all_dynamic_columns = {}  # normalized_key -> original_column_name
        if template:
            existing_template_keys = set(template.fields.values_list('key', flat=True))
        else:
            existing_template_keys = set()

        for h in headers:
            try:
                header_str = str(h).strip()
            except Exception:
                continue
            if not header_str:
                continue
            norm = normalize_header(header_str)
            if not norm:
                continue
            # Ne pas créer si déjà présent dans le template
            if norm in existing_template_keys:
                continue
            # Ne jamais créer de champ pour les colonnes fusionnées
            if norm in _COLUMNS_FUSED_INTO_COMORBIDITE_LISTE:
                continue
            # Si la colonne est mappée dans COLUMN_MAPPING, ce n'est pas une dynamique
            if norm in COLUMN_MAPPING:
                continue
            # Pré-enregistrer comme colonne dynamique (sera marquée/créée plus tard)
            all_dynamic_columns.setdefault(norm, header_str)

        for index, row in dataframe.iterrows():
            payload = build_patient_payload(row)

            # Récupérer les colonnes dynamiques détectées pour cette ligne
            detected = payload.pop('_dynamic_columns_detected', set())
            for norm_key in detected:
                if norm_key not in all_dynamic_columns:
                    # Chercher le nom original dans les headers du dataframe
                    original_name = next(
                        (str(h).strip() for h in headers
                         if normalize_header(str(h).strip()) == norm_key),
                        norm_key,
                    )
                    all_dynamic_columns[norm_key] = original_name

            payload = apply_automatic_schema_fields(
                payload, auto_increment_state=auto_increment_state, current_user=request.user
            )
            payload = ensure_required_identity_fields(payload)
            payload = ensure_incremental_identifiers(
                payload, auto_increment_state=auto_increment_state, force_generated=True
            )

            serializer = PatientSerializer(data=payload)
            if serializer.is_valid():
                serializer.save()
                created_count += 1
            else:
                row_errors.append({'row': index + 2, 'errors': serializer.errors})
                print(f"ERREUR ligne {index + 2}: {serializer.errors}")

        # ── Enregistrer les colonnes dynamiques dans le template ──────────────
        dynamic_columns_info = []
        if all_dynamic_columns and template:
            existing_keys = set(template.fields.values_list('key', flat=True))
            new_fields = []
            for norm_key, original_name in all_dynamic_columns.items():
                if norm_key in existing_keys:
                    # Déjà connue : mettre à jour le source_hint pour marquer comme dynamique
                    template.fields.filter(key=norm_key).update(
                        source_hint='dynamic_column'
                    )
                else:
                    new_fields.append(
                        PatientFormField(
                            template=template,
                            key=norm_key,
                            label=original_name,
                            field_type='text_short',
                            order=10000 + len(new_fields),
                            choices=[],
                            source_hint='dynamic_column',
                            is_required=False,
                        )
                    )
                dynamic_columns_info.append({
                    'key': norm_key,
                    'label': original_name,
                    'is_new': norm_key not in existing_keys,
                })

            if new_fields:
                PatientFormField.objects.bulk_create(new_fields)
                print(f"✅ {len(new_fields)} nouvelle(s) colonne(s) dynamique(s) enregistrée(s): "
                      f"{[f.key for f in new_fields]}")

        if created_count > 0 or all_dynamic_columns:
            try:
                refresh_postgres_flat_view(template)
            except Exception as e:
                print(f"Erreur rafraîchissement vue: {e}")

        template_data = PatientFormTemplateSerializer(template).data
        status_code = status.HTTP_201_CREATED if created_count else status.HTTP_200_OK

        print(f"\n{'='*80}")
        print(f"IMPORT TERMINÉ: {created_count} patients créés, {len(row_errors)} erreurs")
        print(f"Colonnes dynamiques: {list(all_dynamic_columns.keys())}")
        print(f"{'='*80}\n")

        return Response(
            {
                'mode': 'data',
                'template': template_data,
                'fields_created': len(template_data.get('fields', [])),
                'patients_created': created_count,
                'errors': row_errors,
                'dynamic_columns': dynamic_columns_info,
                'dynamic_columns_count': len(dynamic_columns_info),
                'new_dynamic_columns_count': sum(1 for c in dynamic_columns_info if c.get('is_new')),
            },
            status=status_code,
        )


class PatientSchemaView(APIView):
    permission_classes = [CanViewPatients]

    def get(self, request):
        template = get_active_template()
        if not template:
            return Response({'template': None})

        # Supprimer définitivement les colonnes déjà fusionnées dans comorbidite_liste
        template.fields.filter(key__in=_COLUMNS_FUSED_INTO_COMORBIDITE_LISTE).delete()

        return Response({'template': PatientFormTemplateSerializer(template).data})


class PatientPlateformeFlatView(APIView):
    permission_classes = [CanViewPatients]

    def get(self, request):
        with connection.cursor() as cursor:
            cursor.execute('SELECT * FROM patients_plateforme_flat')
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        return Response(rows)